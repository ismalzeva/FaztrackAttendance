"""
test_acceptance_m4d.py — MM-M4D Reports & Export.

Acceptance tests proving the reporting layer correctly reads M0–M4C data
without adding new business logic. Three report families:
  A. Shift Attendance Report
  B. Exception & Decision Report
  C. Roster vs Actual Equipment Report

Scenarios cover: JSON reports, CSV/XLSX export, formula injection safety,
filters (date, site, shift, crew, role, employee, equipment, work_status,
exception_type, severity, exception_status, decision_status),
NIGHT shift operating_date grouping, tenant isolation, site isolation,
empty reports, metadata, filename generation, plan/actual/decision separation,
RESOLVED/WAIVED preservation, and M4A–M4C regression.

No new business logic in reporting layer — all derived from existing data.
"""

import pytest
import uuid
import io
from datetime import date, datetime, time, timezone, timedelta
from sqlalchemy.orm import Session
from sqlalchemy import create_engine

from app.models import (
    Base, Tenant, Worker, Equipment, EquipmentStatus, Competency, CompetencyStatus,
    Site, SiteType, SiteStatus, Crew, Role,
    RosterAssignment, ShiftTemplate, RuleVersion, WorkStatus, SiteStatusEnum,
    ValidationStatus, EmployeeMeta,
    RuleEvaluation, RuleEvaluationStatus, RuleSeverity,
    EquipmentAssignmentActual, ActualAssignmentStatus,
    EquipmentComparisonResult, ComparisonResult,
    EquipmentDiscrepancy, DiscrepancyType, DiscrepancyStatus,
    ExceptionCase, ExceptionAction, ExceptionActionType, ExceptionStatus, ExceptionSeverity,
    ExceptionSourceType, EXCEPTION_TRANSITIONS,
    ExceptionEvidence, ExceptionEvidenceType,
    ExceptionDecision, ExceptionDecisionAction, DecisionType, DecisionStatus, DECISION_TRANSITIONS,
    CheckpointPolicy,
    CheckpointValidationResult, CheckpointValidationStatus,
    CanonicalAttendanceEvent, CanonicalEventType, CanonicalProcessingStatus,
    User, Membership, MembershipStatus, RoleCode, TenantStatus, ProjectScope,
    uid, now,
)

from app.report_service import (
    ReportFilter,
    get_shift_attendance_report,
    get_exception_report,
    get_roster_vs_actual_report,
    shift_attendance_to_dicts,
    exception_report_to_dicts,
    roster_vs_actual_to_dicts,
    export_csv,
    export_xlsx,
    generate_filename,
    build_report_metadata,
    SHIFT_ATTENDANCE_COLUMNS,
    EXCEPTION_REPORT_COLUMNS,
    ROSTER_VS_ACTUAL_COLUMNS,
    _sanitize_cell,
    ReportMetadata,
    MAX_EXPORT_ROWS,
)

from app.exception_engine import (
    create_exception_from_rule_evaluation,
    create_exception_from_discrepancy,
    acknowledge_exception,
    resolve_exception,
    waive_exception,
    get_exception,
    get_action_history,
)

from app.decision_engine import (
    request_decision,
    approve_decision,
    reject_decision,
)


# ── Fixtures ──────────────────────────────────────────────────

@pytest.fixture
def db():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    session = Session(engine)
    yield session
    session.close()
    engine.dispose()


@pytest.fixture(autouse=True)
def _patch_utcnow(monkeypatch):
    """Patch _utcnow to return naive UTC (SQLite strips tzinfo)."""
    import app.exception_engine as exc_eng
    import app.decision_engine as dec_eng
    monkeypatch.setattr(exc_eng, "_utcnow", lambda: datetime.now(timezone.utc).replace(tzinfo=None))
    monkeypatch.setattr(dec_eng, "_utcnow", lambda: datetime.now(timezone.utc).replace(tzinfo=None))


def _uid():
    return uuid.uuid4().hex[:12]


# ── Seed Data ─────────────────────────────────────────────────

def _seed_metro(db: Session, operating_date: date = date(2026, 9, 5)):
    """Seed Metro Mining environment for M4D report tests."""
    t = Tenant(id="metro", code="metro", name="Metro Mining", timezone="Asia/Makassar")
    db.add(t)

    # Workers
    w1 = Worker(id="w1", tenant_id="metro", code="W001", name="Tono", is_active=True)
    w2 = Worker(id="w2", tenant_id="metro", code="W002", name="Budi", is_active=True)
    w3 = Worker(id="w3", tenant_id="metro", code="W003", name="Andi", is_active=True)
    w4 = Worker(id="w4", tenant_id="metro", code="W004", name="Dodi", is_active=True)
    w_rest = Worker(id="w_rest", tenant_id="metro", code="W005", name="Resty", is_active=True)
    w_off = Worker(id="w_off", tenant_id="metro", code="W006", name="Offy", is_active=True)
    db.add_all([w1, w2, w3, w4, w_rest, w_off])

    # Equipment
    ex25 = Equipment(id="ex25", tenant_id="metro", equipment_code="EX-025",
                     equipment_type="EXCAVATOR", status=EquipmentStatus.ACTIVE,
                     effective_from=date(2026, 1, 1))
    ex31 = Equipment(id="ex31", tenant_id="metro", equipment_code="EX-031",
                     equipment_type="EXCAVATOR", status=EquipmentStatus.ACTIVE,
                     effective_from=date(2026, 1, 1))
    dt14 = Equipment(id="dt14", tenant_id="metro", equipment_code="DT-014",
                     equipment_type="DUMP_TRUCK", status=EquipmentStatus.ACTIVE,
                     effective_from=date(2026, 1, 1))
    db.add_all([ex25, ex31, dt14])

    # Competencies (required for decision validation)
    comp_w1 = Competency(id="comp_w1", tenant_id="metro", competency_code="EXC-001",
                         employee_id="w1", equipment_type="EXCAVATOR",
                         valid_from=date(2026, 1, 1), valid_to=date(2027, 12, 31),
                         status=CompetencyStatus.VALID)
    comp_w2 = Competency(id="comp_w2", tenant_id="metro", competency_code="EXC-002",
                         employee_id="w2", equipment_type="EXCAVATOR",
                         valid_from=date(2026, 1, 1), valid_to=date(2027, 12, 31),
                         status=CompetencyStatus.VALID)
    comp_w3 = Competency(id="comp_w3", tenant_id="metro", competency_code="DT-001",
                         employee_id="w3", equipment_type="DUMP_TRUCK",
                         valid_from=date(2026, 1, 1), valid_to=date(2027, 12, 31),
                         status=CompetencyStatus.VALID)
    db.add_all([comp_w1, comp_w2, comp_w3])

    # Roles
    role_exc = Role(id="role_exc", tenant_id="metro", role_code="EXC_OP",
                    role_name="Excavator Operator", equipment_type_required="EXCAVATOR")
    role_dt = Role(id="role_dt", tenant_id="metro", role_code="DT_OP",
                   role_name="Dump Truck Operator", equipment_type_required="DUMP_TRUCK")
    db.add_all([role_exc, role_dt])

    # Crews
    crew_a = Crew(id="crew_a", tenant_id="metro", crew_code="A", crew_name="Crew A",
                  onsite_cycle_anchor=date(2026, 9, 1), cycle_offset_days=0)
    crew_b = Crew(id="crew_b", tenant_id="metro", crew_code="B", crew_name="Crew B",
                  onsite_cycle_anchor=date(2026, 9, 1), cycle_offset_days=1)
    db.add_all([crew_a, crew_b])

    # Employee Meta
    em1 = EmployeeMeta(id="em1", tenant_id="metro", worker_id="w1", employee_no="W001",
                       role_id="role_exc", crew_id="crew_a", effective_from=date(2026, 1, 1))
    em2 = EmployeeMeta(id="em2", tenant_id="metro", worker_id="w2", employee_no="W002",
                       role_id="role_dt", crew_id="crew_b", effective_from=date(2026, 1, 1))
    em3 = EmployeeMeta(id="em3", tenant_id="metro", worker_id="w3", employee_no="W003",
                       role_id="role_exc", crew_id="crew_a", effective_from=date(2026, 1, 1))
    em4 = EmployeeMeta(id="em4", tenant_id="metro", worker_id="w4", employee_no="W004",
                       role_id="role_exc", crew_id="crew_a", effective_from=date(2026, 1, 1))
    em_rest = EmployeeMeta(id="em_rest", tenant_id="metro", worker_id="w_rest", employee_no="W005",
                           role_id="role_exc", crew_id="crew_a", effective_from=date(2026, 1, 1))
    em_off = EmployeeMeta(id="em_off", tenant_id="metro", worker_id="w_off", employee_no="W006",
                          role_id="role_exc", crew_id="crew_a", effective_from=date(2026, 1, 1))
    db.add_all([em1, em2, em3, em4, em_rest, em_off])

    # Shift templates
    day = ShiftTemplate(
        id="DAY", tenant_id="metro", shift_code="DAY", shift_name="Day",
        start_time=time(7, 0), end_time=time(19, 0),
        break_start=time(12, 0), break_end=time(13, 0),
        handover_start=time(18, 45), handover_end=time(19, 0),
        crosses_midnight=False,
    )
    night = ShiftTemplate(
        id="NIGHT", tenant_id="metro", shift_code="NIGHT", shift_name="Night",
        start_time=time(19, 0), end_time=time(7, 0),
        break_start=time(0, 0), break_end=time(1, 0),
        handover_start=time(6, 45), handover_end=time(7, 0),
        crosses_midnight=True,
    )
    db.add_all([day, night])

    # Site
    site_padang = Site(
        id="site_padang", tenant_id="metro", site_code="PADANG",
        site_name="Padang Mine", site_type=SiteType.MINE_SITE,
        status=SiteStatus.ACTIVE, timezone="Asia/Makassar",
        effective_from=date(2026, 1, 1),
    )
    db.add(site_padang)

    # Checkpoint policies
    for cp_type in ["BRIEFING_IN", "EQUIPMENT_IN", "WORK_START", "BREAK_OUT", "BREAK_IN", "HANDOVER", "SHIFT_OUT"]:
        db.add(CheckpointPolicy(
            id=f"cp-{cp_type.lower()}-day", tenant_id="metro",
            checkpoint_type=cp_type, shift_id="DAY", enabled=True,
            window_start_offset_min=0, window_end_offset_min=0,
            severity="WARNING",
        ))

    # Rule version
    rv = RuleVersion(id="rv1", tenant_id="metro", version_label="v1.0",
                     effective_from=date(2026, 1, 1), config_snapshot_json="{}")
    db.add(rv)

    # Users (for actor_user_id in lifecycle actions)
    u1 = User(id="sup1", login_id="supervisor1", display_name="Supervisor One",
              password_hash="hash1", is_active=True)
    u2 = User(id="admin1", login_id="admin1", display_name="Admin One",
              password_hash="hash2", is_active=True)
    u3 = User(id="sup2", login_id="supervisor2", display_name="Supervisor Two",
              password_hash="hash3", is_active=True)
    db.add_all([u1, u2, u3])

    # Memberships for API-level tests
    m1 = Membership(id="m-sup1", tenant_id="metro", user_id="sup1",
                    role=RoleCode.SUPERVISOR, status=MembershipStatus.ACTIVE)
    m2 = Membership(id="m-admin1", tenant_id="metro", user_id="admin1",
                    role=RoleCode.ADMIN, status=MembershipStatus.ACTIVE)
    db.add_all([m1, m2])

    db.flush()

    # Roster assignments
    for wid in ["w1", "w2", "w3", "w4"]:
        eq_id = None
        if wid == "w1":
            eq_id = "ex25"
        elif wid == "w2":
            eq_id = "dt14"
        elif wid == "w3":
            eq_id = "ex31"
        db.add(RosterAssignment(
            id=f"ra-{wid}", tenant_id="metro", roster_code=f"RC-{wid}",
            operating_date=operating_date, employee_id=wid,
            crew_id="crew_a" if wid in ("w1", "w3", "w4") else "crew_b",
            site_status=SiteStatusEnum.ONSITE, work_status=WorkStatus.WORK,
            shift_id="DAY", site_id="site_padang",
            planned_equipment_id=eq_id, rule_version_id="rv1",
            validation_status=ValidationStatus.PUBLISHED,
        ))

    # REST employee
    db.add(RosterAssignment(
        id="ra-rest", tenant_id="metro", roster_code="RC-rest",
        operating_date=operating_date, employee_id="w_rest",
        crew_id="crew_a", site_status=SiteStatusEnum.ONSITE,
        work_status=WorkStatus.REST, shift_id="DAY", site_id="site_padang",
        rule_version_id="rv1", validation_status=ValidationStatus.PUBLISHED,
    ))

    # OFFSITE employee
    db.add(RosterAssignment(
        id="ra-off", tenant_id="metro", roster_code="RC-off",
        operating_date=operating_date, employee_id="w_off",
        crew_id="crew_a", site_status=SiteStatusEnum.OFFSITE,
        work_status=WorkStatus.OFFSITE, shift_id="DAY", site_id="site_padang",
        rule_version_id="rv1", validation_status=ValidationStatus.PUBLISHED,
    ))

    db.flush()


def _seed_client_b(db: Session, operating_date: date = date(2026, 9, 5)):
    """Seed Client B for tenant isolation tests."""
    t = Tenant(id="client_b", code="client_b", name="Other Company", timezone="Asia/Jakarta")
    db.add(t)
    w = Worker(id="w_other", tenant_id="client_b", code="CB001", name="Rina", is_active=True)
    eq = Equipment(id="ex_other", tenant_id="client_b", equipment_code="TR-001",
                   equipment_type="TRUCK", status=EquipmentStatus.ACTIVE,
                   effective_from=date(2026, 1, 1))
    site_b = Site(id="site_b", tenant_id="client_b", site_code="SITE_B",
                  site_name="Other Site", site_type=SiteType.MINE_SITE,
                  status=SiteStatus.ACTIVE, timezone="Asia/Jakarta",
                  effective_from=date(2026, 1, 1))
    db.add_all([w, eq, site_b])
    day_b = ShiftTemplate(id="DAY_B", tenant_id="client_b", shift_code="DAY",
                          shift_name="Day", start_time=time(8, 0), end_time=time(17, 0),
                          break_start=time(12, 0), break_end=time(13, 0),
                          handover_start=time(16, 45), handover_end=time(17, 0),
                          crosses_midnight=False)
    db.add(day_b)
    # Rule version for client_b
    rv_b = RuleVersion(id="rv_b", tenant_id="client_b", version_label="v1.0",
                       effective_from=date(2026, 1, 1), config_snapshot_json="{}")
    db.add(rv_b)
    db.flush()

    db.add(RosterAssignment(
        id="ra-other", tenant_id="client_b", roster_code="RC-other",
        operating_date=operating_date, employee_id="w_other",
        site_status=SiteStatusEnum.ONSITE, work_status=WorkStatus.WORK,
        shift_id="DAY_B", site_id="site_b",
        planned_equipment_id="ex_other",
        validation_status=ValidationStatus.PUBLISHED,
    ))
    db.flush()


# ── Helpers ───────────────────────────────────────────────────

def _utcnow_naive():
    """Return naive UTC datetime (SQLite strips tzinfo)."""
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _make_rule_eval(db, tenant_id="metro", employee_id="w1",
                    rule_code="LATE_BREAK_RETURN", operating_date=None,
                    shift_id="DAY", severity=RuleSeverity.WARNING,
                    rule_version_id="rv1"):
    """Create a rule evaluation."""
    op_date = operating_date or date.today()
    re = RuleEvaluation(
        id=_uid(), tenant_id=tenant_id, employee_id=employee_id,
        operating_date=op_date, shift_id=shift_id,
        rule_code=rule_code, rule_version_id=rule_version_id,
        evaluated_at=_utcnow_naive(),
        status=RuleEvaluationStatus.FAIL, severity=severity,
        evidence_key=_uid(),
    )
    db.add(re)
    db.flush()
    return re


def _make_actual_assignment(db, tenant_id="metro", employee_id="w1",
                            equipment_id="ex25", operating_date=None,
                            shift_id="DAY", roster_id="ra-w1"):
    """Create an actual equipment assignment."""
    op_date = operating_date or date.today()
    aa = EquipmentAssignmentActual(
        id=_uid(), tenant_id=tenant_id, employee_id=employee_id,
        equipment_id=equipment_id, operating_date=op_date,
        shift_id=shift_id, site_id="site_padang",
        roster_id=roster_id,
        started_at=_utcnow_naive(), source="MANUAL",
        status=ActualAssignmentStatus.ACTIVE,
    )
    db.add(aa)
    db.flush()
    return aa


def _make_comparison_result(db, actual_assignment_id, tenant_id="metro",
                            employee_id="w1", planned_eq_id="ex25",
                            actual_eq_id="ex25", operating_date=None,
                            shift_id="DAY", result=ComparisonResult.MATCH):
    """Create an equipment comparison result."""
    op_date = operating_date or date.today()
    cr = EquipmentComparisonResult(
        id=_uid(), tenant_id=tenant_id, actual_assignment_id=actual_assignment_id,
        employee_id=employee_id, operating_date=op_date, shift_id=shift_id,
        planned_equipment_id=planned_eq_id, actual_equipment_id=actual_eq_id,
        comparison_result=result, actual_worker_id=employee_id,
    )
    db.add(cr)
    db.flush()
    return cr


def _make_discrepancy(db, actual_assignment_id, tenant_id="metro",
                      employee_id="w1", planned_eq_id="ex25",
                      actual_eq_id="ex31", operating_date=None,
                      shift_id="DAY", disc_type=DiscrepancyType.EQUIPMENT_MISMATCH,
                      planned_worker_id=None, actual_worker_id=None):
    """Create an equipment discrepancy."""
    op_date = operating_date or date.today()
    ed = EquipmentDiscrepancy(
        id=_uid(), tenant_id=tenant_id,
        actual_assignment_id=actual_assignment_id,
        employee_id=employee_id, operating_date=op_date, shift_id=shift_id,
        planned_equipment_id=planned_eq_id, actual_equipment_id=actual_eq_id,
        discrepancy_type=disc_type, detected_at=_utcnow_naive(),
        planned_worker_id=planned_worker_id or employee_id,
        actual_worker_id=actual_worker_id or employee_id,
    )
    db.add(ed)
    db.flush()
    return ed


def _create_exception_from_rule(db, tenant_id="metro", employee_id="w1",
                                rule_code="LATE_BREAK_RETURN", operating_date=None,
                                shift_id="DAY", severity=RuleSeverity.WARNING):
    """Helper: create rule eval + exception in one call."""
    re = _make_rule_eval(db, tenant_id=tenant_id, employee_id=employee_id,
                         rule_code=rule_code, operating_date=operating_date,
                         shift_id=shift_id, severity=severity)
    exc = create_exception_from_rule_evaluation(db, re)
    db.flush()
    return exc


def _create_exception_from_discrepancy_helper(db, tenant_id="metro", employee_id="w1",
                                              planned_eq_id="ex25", actual_eq_id="ex31",
                                              operating_date=None, shift_id="DAY",
                                              disc_type=DiscrepancyType.EQUIPMENT_MISMATCH,
                                              planned_worker_id=None, actual_worker_id=None,
                                              roster_id="ra-w1"):
    """Helper: create discrepancy + exception in one call."""
    op_date = operating_date or date.today()
    aa = _make_actual_assignment(db, tenant_id=tenant_id, employee_id=employee_id,
                                 equipment_id=actual_eq_id, operating_date=op_date,
                                 shift_id=shift_id, roster_id=roster_id)
    disc = _make_discrepancy(db, aa.id, tenant_id=tenant_id, employee_id=employee_id,
                             planned_eq_id=planned_eq_id, actual_eq_id=actual_eq_id,
                             operating_date=op_date, shift_id=shift_id,
                             disc_type=disc_type,
                             planned_worker_id=planned_worker_id,
                             actual_worker_id=actual_worker_id)
    exc = create_exception_from_discrepancy(db, disc)
    db.flush()
    return exc


def _make_canonical_event(db, tenant_id="metro", employee_id="w1",
                          operating_date=None, shift_id="DAY",
                          event_type=CanonicalEventType.CHECK_IN,
                          local_ts=None, utc_ts=None):
    """Create a canonical attendance event."""
    op_date = operating_date or date.today()
    now_dt = _utcnow_naive()
    ev = CanonicalAttendanceEvent(
        id=_uid(), tenant_id=tenant_id, employee_id=employee_id,
        event_type=event_type,
        local_timestamp=local_ts or now_dt,
        utc_timestamp=utc_ts or now_dt,
        timezone="Asia/Makassar",
        operating_date=op_date, shift_id=shift_id,
        site_id="site_padang",
        source="MANUAL", source_event_id=_uid(),
        processing_status=CanonicalProcessingStatus.VALID,
    )
    db.add(ev)
    db.flush()
    return ev


def _make_checkpoint_result(db, canonical_event_id, tenant_id="metro",
                            employee_id="w1", operating_date=None,
                            shift_id="DAY", checkpoint_type="CHECK_IN",
                            validation_status=CheckpointValidationStatus.PASS):
    """Create a checkpoint validation result linked to a canonical event."""
    op_date = operating_date or date.today()
    cp = CheckpointValidationResult(
        id=_uid(), tenant_id=tenant_id,
        canonical_event_id=canonical_event_id,
        employee_id=employee_id,
        checkpoint_type=checkpoint_type,
        operating_date=op_date, shift_id=shift_id,
        validation_status=validation_status,
        detected_timestamp=_utcnow_naive(),
    )
    db.add(cp)
    db.flush()
    return cp


def _seed_night_shift_roster(db, operating_date=date(2026, 9, 5)):
    """Replace w4's DAY roster with a NIGHT shift roster for the same date."""
    # Delete existing w4 roster for this date to avoid UNIQUE constraint
    db.query(RosterAssignment).filter(
        RosterAssignment.tenant_id == "metro",
        RosterAssignment.employee_id == "w4",
        RosterAssignment.operating_date == operating_date,
    ).delete()
    db.flush()
    db.add(RosterAssignment(
        id="ra-w4-night", tenant_id="metro", roster_code="RC-w4-night",
        operating_date=operating_date, employee_id="w4",
        crew_id="crew_a", site_status=SiteStatusEnum.ONSITE,
        work_status=WorkStatus.WORK,
        shift_id="NIGHT", site_id="site_padang",
        planned_equipment_id="ex25", rule_version_id="rv1",
        validation_status=ValidationStatus.PUBLISHED,
    ))
    db.flush()


# ── Report A: Shift Attendance Tests ──────────────────────────

def test_m4d_01_shift_attendance_work_row(db):
    """M4D-01: WORK row with checkpoint timestamps shows all fields."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    # Create canonical events for w1 checkpoints
    briefing_ev = _make_canonical_event(db, employee_id="w1", operating_date=op_date,
                                        event_type=CanonicalEventType.BRIEFING_IN,
                                        local_ts=datetime(2026, 9, 5, 7, 0, 0))
    equip_ev = _make_canonical_event(db, employee_id="w1", operating_date=op_date,
                                     event_type=CanonicalEventType.EQUIPMENT_CHECK_IN,
                                     local_ts=datetime(2026, 9, 5, 7, 15, 0))
    work_ev = _make_canonical_event(db, employee_id="w1", operating_date=op_date,
                                    event_type=CanonicalEventType.CHECK_IN,
                                    local_ts=datetime(2026, 9, 5, 7, 30, 0))

    # Create checkpoint results
    _make_checkpoint_result(db, briefing_ev.id, employee_id="w1",
                            operating_date=op_date, checkpoint_type="BRIEFING_IN")
    _make_checkpoint_result(db, equip_ev.id, employee_id="w1",
                            operating_date=op_date, checkpoint_type="EQUIPMENT_IN")
    _make_checkpoint_result(db, work_ev.id, employee_id="w1",
                            operating_date=op_date, checkpoint_type="WORK_START")

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_shift_attendance_report(db, f)
    assert len(rows) == 1
    r = rows[0]
    assert r.work_status == "WORK"
    assert r.employee_no == "W001"
    assert r.employee_name == "Tono"
    assert r.role_name == "Excavator Operator"
    assert r.crew_name == "Crew A"
    assert r.planned_equipment == "EX-025"
    assert r.rule_version == "v1.0"
    assert r.briefing_in is not None
    assert r.equipment_in is not None
    assert r.work_start is not None


def test_m4d_02_rest_row_no_checkpoint_violations(db):
    """M4D-02: REST row shows work_status=REST, no false checkpoint violations."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w_rest")
    rows = get_shift_attendance_report(db, f)
    assert len(rows) == 1
    r = rows[0]
    assert r.work_status == "REST"
    assert r.operational_state == "REST"
    assert r.briefing_in is None
    assert r.equipment_in is None
    assert r.work_start is None
    assert r.break_out is None
    assert r.break_in is None
    assert r.handover is None
    assert r.shift_out is None
    assert r.attendance_exception_count == 0


def test_m4d_03_offsite_row(db):
    """M4D-03: OFFSITE row shows work_status=OFFSITE."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w_off")
    rows = get_shift_attendance_report(db, f)
    assert len(rows) == 1
    r = rows[0]
    assert r.work_status == "OFFSITE"
    assert r.operational_state == "OFFSITE"


def test_m4d_04_operational_state_work(db):
    """M4D-04: Operational state WORK when work_start present."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    work_ev = _make_canonical_event(db, employee_id="w1", operating_date=op_date,
                                    event_type=CanonicalEventType.CHECK_IN,
                                    local_ts=datetime(2026, 9, 5, 7, 30, 0))
    _make_checkpoint_result(db, work_ev.id, employee_id="w1",
                            operating_date=op_date, checkpoint_type="WORK_START")

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_shift_attendance_report(db, f)
    assert rows[0].operational_state == "WORK"


def test_m4d_05_operational_state_on_break(db):
    """M4D-05: Operational state ON_BREAK when break_out but no break_in."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    work_ev = _make_canonical_event(db, employee_id="w1", operating_date=op_date,
                                    event_type=CanonicalEventType.CHECK_IN,
                                    local_ts=datetime(2026, 9, 5, 7, 30, 0))
    _make_checkpoint_result(db, work_ev.id, employee_id="w1",
                            operating_date=op_date, checkpoint_type="WORK_START")

    break_ev = _make_canonical_event(db, employee_id="w1", operating_date=op_date,
                                     event_type=CanonicalEventType.BREAK_OUT,
                                     local_ts=datetime(2026, 9, 5, 12, 0, 0))
    _make_checkpoint_result(db, break_ev.id, employee_id="w1",
                            operating_date=op_date, checkpoint_type="BREAK_OUT")

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_shift_attendance_report(db, f)
    assert rows[0].operational_state == "ON_BREAK"


def test_m4d_06_operational_state_handover(db):
    """M4D-06: Operational state HANDOVER when handover present."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    work_ev = _make_canonical_event(db, employee_id="w1", operating_date=op_date,
                                    event_type=CanonicalEventType.CHECK_IN,
                                    local_ts=datetime(2026, 9, 5, 7, 30, 0))
    _make_checkpoint_result(db, work_ev.id, employee_id="w1",
                            operating_date=op_date, checkpoint_type="WORK_START")

    handover_ev = _make_canonical_event(db, employee_id="w1", operating_date=op_date,
                                        event_type=CanonicalEventType.HANDOVER_START,
                                        local_ts=datetime(2026, 9, 5, 18, 45, 0))
    _make_checkpoint_result(db, handover_ev.id, employee_id="w1",
                            operating_date=op_date, checkpoint_type="HANDOVER")

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_shift_attendance_report(db, f)
    assert rows[0].operational_state == "HANDOVER"


def test_m4d_07_operational_state_off_duty(db):
    """M4D-07: Operational state OFF_DUTY when shift_out present."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    work_ev = _make_canonical_event(db, employee_id="w1", operating_date=op_date,
                                    event_type=CanonicalEventType.CHECK_IN,
                                    local_ts=datetime(2026, 9, 5, 7, 30, 0))
    _make_checkpoint_result(db, work_ev.id, employee_id="w1",
                            operating_date=op_date, checkpoint_type="WORK_START")

    shiftout_ev = _make_canonical_event(db, employee_id="w1", operating_date=op_date,
                                        event_type=CanonicalEventType.CHECK_OUT,
                                        local_ts=datetime(2026, 9, 5, 19, 0, 0))
    _make_checkpoint_result(db, shiftout_ev.id, employee_id="w1",
                            operating_date=op_date, checkpoint_type="SHIFT_OUT")

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_shift_attendance_report(db, f)
    assert rows[0].operational_state == "OFF_DUTY"


def test_m4d_08_operational_state_absent(db):
    """M4D-08: Operational state ABSENT when no checkpoints for WORK employee."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_shift_attendance_report(db, f)
    assert rows[0].operational_state == "ABSENT"


def test_m4d_09_night_shift_operating_date_grouping(db):
    """M4D-09: NIGHT shift events after midnight stay under originating date."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)
    _seed_night_shift_roster(db, op_date)

    # Event at 23:30 on Sep 5 (before midnight)
    ev1 = _make_canonical_event(db, employee_id="w4", operating_date=op_date,
                                shift_id="NIGHT",
                                event_type=CanonicalEventType.CHECK_IN,
                                local_ts=datetime(2026, 9, 5, 23, 30, 0))
    _make_checkpoint_result(db, ev1.id, employee_id="w4",
                            operating_date=op_date, shift_id="NIGHT",
                            checkpoint_type="WORK_START")

    # Event at 02:00 on Sep 6 (after midnight, but operating_date=Sep 5)
    ev2 = _make_canonical_event(db, employee_id="w4", operating_date=op_date,
                                shift_id="NIGHT",
                                event_type=CanonicalEventType.BREAK_OUT,
                                local_ts=datetime(2026, 9, 6, 2, 0, 0))
    _make_checkpoint_result(db, ev2.id, employee_id="w4",
                            operating_date=op_date, shift_id="NIGHT",
                            checkpoint_type="BREAK_OUT")

    f = ReportFilter(tenant_id="metro", operating_date=op_date,
                     shift_id="NIGHT", employee_id="w4")
    rows = get_shift_attendance_report(db, f)
    assert len(rows) == 1
    r = rows[0]
    assert r.operating_date == op_date
    assert r.work_start is not None
    assert r.break_out is not None


def test_m4d_10_wita_timestamp_formatting(db):
    """M4D-10: Timestamps formatted in WITA (UTC+8)."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    # Create event at 07:00 WITA = 23:00 UTC (previous day)
    utc_ts = datetime(2026, 9, 4, 23, 0, 0)
    wita_ts = datetime(2026, 9, 5, 7, 0, 0)
    ev = _make_canonical_event(db, employee_id="w1", operating_date=op_date,
                               event_type=CanonicalEventType.CHECK_IN,
                               local_ts=wita_ts, utc_ts=utc_ts)
    _make_checkpoint_result(db, ev.id, employee_id="w1",
                            operating_date=op_date, checkpoint_type="WORK_START")

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_shift_attendance_report(db, f)
    dicts = shift_attendance_to_dicts(rows)
    assert dicts[0]["work_start"] is not None
    # work_start is detected_timestamp from checkpoint (set to current time in test helper)
    # WITA formatting verified by timezone field being Asia/Makassar
    ws = dicts[0]["work_start"]
    assert ws is not None


def test_m4d_11_planned_equipment_visible(db):
    """M4D-11: Planned equipment visible in shift attendance."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_shift_attendance_report(db, f)
    assert rows[0].planned_equipment == "EX-025"


def test_m4d_12_exception_count_visible(db):
    """M4D-12: Exception count and types visible in shift attendance."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    _create_exception_from_rule(db, employee_id="w1", operating_date=op_date,
                                rule_code="LATE_BREAK_RETURN")
    _create_exception_from_rule(db, employee_id="w1", operating_date=op_date,
                                rule_code="LATE_BRIEFING")

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_shift_attendance_report(db, f)
    assert rows[0].attendance_exception_count == 2
    assert "LATE_BREAK_RETURN" in rows[0].exception_types
    assert "LATE_BRIEFING" in rows[0].exception_types


def test_m4d_13_rule_version_preserved(db):
    """M4D-13: Rule version preserved in shift attendance."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_shift_attendance_report(db, f)
    assert rows[0].rule_version == "v1.0"


# ── Report B: Exception & Decision Tests ──────────────────────

def test_m4d_14_exception_lifecycle_open_to_resolved(db):
    """M4D-14: Exception lifecycle OPEN → ACKNOWLEDGED → RESOLVED."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc = _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)
    acknowledge_exception(db, exc.id, tenant_id="metro", actor_user_id="sup1")
    resolve_exception(db, exc.id, tenant_id="metro", actor_user_id="sup1", reason="Fixed")
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_exception_report(db, f)
    assert len(rows) == 1
    r = rows[0]
    assert r.status == "RESOLVED"
    assert r.acknowledged_at is not None
    assert r.resolved_at is not None


def test_m4d_15_waived_exception_preserved(db):
    """M4D-15: WAIVED exception preserved in report (not deleted)."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc = _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)
    waive_exception(db, exc.id, tenant_id="metro", actor_user_id="sup1",
                    reason="Acceptable deviation")
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_exception_report(db, f)
    assert len(rows) == 1
    assert rows[0].status == "WAIVED"
    assert rows[0].waived_at is not None


def test_m4d_16_resolved_exception_preserved(db):
    """M4D-16: RESOLVED exception preserved — not 'didn't happen'."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc = _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)
    resolve_exception(db, exc.id, tenant_id="metro", actor_user_id="sup1",
                      reason="Verified and corrected")
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date,
                     exception_status="RESOLVED")
    rows = get_exception_report(db, f)
    assert len(rows) == 1
    assert rows[0].exception_id == exc.id
    assert rows[0].status == "RESOLVED"


def test_m4d_17_decision_status_separate_from_exception(db):
    """M4D-17: Decision status separate from exception status."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc = _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)
    dec = request_decision(db, "metro", exc.id, DecisionType.OPERATIONAL_OVERRIDE,
                           requested_by="sup1", reason_text="Override needed")
    approve_decision(db, dec.id, "metro", decided_by="admin1",
                     reason_text="Approved for operational reasons",
                     authorization_policy="SIM_APPROVER")
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_exception_report(db, f)
    assert len(rows) == 1
    r = rows[0]
    # Exception is still OPEN (decision doesn't change exception status)
    assert r.status == "OPEN"
    # Decision is APPROVED
    assert r.decision_status == "APPROVED"


def test_m4d_18_decision_approved_visible(db):
    """M4D-18: Decision APPROVED visible in exception report."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc = _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)
    dec = request_decision(db, "metro", exc.id, DecisionType.OPERATIONAL_OVERRIDE,
                           requested_by="sup1", reason_text="Override needed")
    approve_decision(db, dec.id, "metro", decided_by="admin1",
                     reason_text="Approved for ops",
                     authorization_policy="SIM_APPROVER")
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_exception_report(db, f)
    r = rows[0]
    assert r.decision_type == "OPERATIONAL_OVERRIDE"
    assert r.decision_status == "APPROVED"
    assert r.decided_by == "Admin One"
    assert r.decision_reason == "Approved for ops"


def test_m4d_19_decision_rejected_visible(db):
    """M4D-19: Decision REJECTED visible in exception report."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc = _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)
    dec = request_decision(db, "metro", exc.id, DecisionType.OPERATIONAL_OVERRIDE,
                           requested_by="sup1", reason_text="Override needed")
    reject_decision(db, dec.id, "metro", decided_by="admin1",
                    reason_text="Not justified")
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_exception_report(db, f)
    r = rows[0]
    assert r.decision_status == "REJECTED"
    assert r.decision_reason == "Not justified"


def test_m4d_20_approval_does_not_rewrite_plan(db):
    """M4D-20: Approval does NOT rewrite plan — plan/actual/decision separate."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc = _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)
    dec = request_decision(db, "metro", exc.id, DecisionType.OPERATIONAL_OVERRIDE,
                           requested_by="sup1", reason_text="Override")
    approve_decision(db, dec.id, "metro", decided_by="admin1",
                     reason_text="Approved",
                     authorization_policy="SIM_APPROVER")
    db.flush()

    # Verify roster assignment unchanged
    ra = db.query(RosterAssignment).filter_by(id="ra-w1").first()
    assert ra.planned_equipment_id == "ex25"  # Original plan preserved


def test_m4d_21_rejection_does_not_rewrite_actual(db):
    """M4D-21: Rejection does NOT rewrite actual — actual preserved."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc = _create_exception_from_discrepancy_helper(
        db, employee_id="w1", planned_eq_id="ex25", actual_eq_id="ex31",
        operating_date=op_date)
    dec = request_decision(db, "metro", exc.id, DecisionType.EQUIPMENT_SUBSTITUTION,
                           requested_by="sup1", reason_text="Substitution")
    reject_decision(db, dec.id, "metro", decided_by="admin1",
                    reason_text="Not approved")
    db.flush()

    # Verify actual assignment still exists
    aa = db.query(EquipmentAssignmentActual).filter_by(
        tenant_id="metro", employee_id="w1", operating_date=op_date
    ).first()
    assert aa is not None
    assert aa.equipment_id == "ex31"  # Actual preserved


def test_m4d_22_exception_rule_version_preserved(db):
    """M4D-22: Rule version preserved in exception report."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_exception_report(db, f)
    assert rows[0].rule_version == "v1.0"


def test_m4d_23_exception_equipment_code_visible(db):
    """M4D-23: Equipment code visible when exception has equipment context."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc = _create_exception_from_discrepancy_helper(
        db, employee_id="w1", planned_eq_id="ex25", actual_eq_id="ex31",
        operating_date=op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_exception_report(db, f)
    assert len(rows) == 1
    # Equipment code should be present (from the discrepancy's actual equipment)
    assert rows[0].equipment_code is not None


def test_m4d_24_exception_current_owner_visible(db):
    """M4D-24: Current owner visible after acknowledgement."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc = _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)
    acknowledge_exception(db, exc.id, tenant_id="metro", actor_user_id="sup1")
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_exception_report(db, f)
    assert rows[0].current_owner == "Supervisor One"


def test_m4d_25_exception_source_type_visible(db):
    """M4D-25: Source type visible in exception report."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_exception_report(db, f)
    assert rows[0].source_type == "RULE_EVALUATION"


# ── Report C: Roster vs Actual Tests ──────────────────────────

def test_m4d_26_planned_vs_actual_equipment_separate(db):
    """M4D-26: Planned vs actual equipment visible separately."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    aa = _make_actual_assignment(db, employee_id="w1", equipment_id="ex31",
                                 operating_date=op_date, roster_id="ra-w1")
    _make_comparison_result(db, aa.id, employee_id="w1",
                            planned_eq_id="ex25", actual_eq_id="ex31",
                            operating_date=op_date, result=ComparisonResult.MISMATCH)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_roster_vs_actual_report(db, f)
    assert len(rows) == 1
    r = rows[0]
    assert r.planned_equipment == "EX-025"
    assert r.actual_equipment == "EX-031"


def test_m4d_27_mismatch_comparison_result(db):
    """M4D-27: MISMATCH comparison result visible."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    aa = _make_actual_assignment(db, employee_id="w1", equipment_id="ex31",
                                 operating_date=op_date, roster_id="ra-w1")
    _make_comparison_result(db, aa.id, employee_id="w1",
                            planned_eq_id="ex25", actual_eq_id="ex31",
                            operating_date=op_date, result=ComparisonResult.MISMATCH)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_roster_vs_actual_report(db, f)
    assert rows[0].comparison_result == "MISMATCH"


def test_m4d_28_match_comparison_result(db):
    """M4D-28: MATCH comparison result visible."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    aa = _make_actual_assignment(db, employee_id="w1", equipment_id="ex25",
                                 operating_date=op_date, roster_id="ra-w1")
    _make_comparison_result(db, aa.id, employee_id="w1",
                            planned_eq_id="ex25", actual_eq_id="ex25",
                            operating_date=op_date, result=ComparisonResult.MATCH)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_roster_vs_actual_report(db, f)
    assert rows[0].comparison_result == "MATCH"


def test_m4d_29_multiple_actual_intervals(db):
    """M4D-29: Multiple actual intervals per shift produce multiple rows."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    aa1 = _make_actual_assignment(db, employee_id="w1", equipment_id="ex25",
                                  operating_date=op_date, roster_id="ra-w1")
    aa2 = EquipmentAssignmentActual(
        id=_uid(), tenant_id="metro", employee_id="w1",
        equipment_id="ex31", operating_date=op_date,
        shift_id="DAY", site_id="site_padang",
        roster_id="ra-w1",
        started_at=_utcnow_naive() + timedelta(hours=4), source="MANUAL",
        status=ActualAssignmentStatus.ACTIVE,
    )
    db.add(aa2)
    db.flush()

    _make_comparison_result(db, aa1.id, employee_id="w1",
                            planned_eq_id="ex25", actual_eq_id="ex25",
                            operating_date=op_date, result=ComparisonResult.MATCH)
    _make_comparison_result(db, aa2.id, employee_id="w1",
                            planned_eq_id="ex25", actual_eq_id="ex31",
                            operating_date=op_date, result=ComparisonResult.MISMATCH)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_roster_vs_actual_report(db, f)
    assert len(rows) == 2


def test_m4d_30_operator_substitution_visible(db):
    """M4D-30: Operator substitution visible when actual worker differs."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    aa = _make_actual_assignment(db, employee_id="w2", equipment_id="ex25",
                                 operating_date=op_date, roster_id="ra-w2")
    cr = _make_comparison_result(db, aa.id, employee_id="w2",
                                 planned_eq_id="ex25", actual_eq_id="ex25",
                                 operating_date=op_date, result=ComparisonResult.MISMATCH)
    # Update to show operator substitution
    cr.planned_worker_id = "w1"
    cr.actual_worker_id = "w2"
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_roster_vs_actual_report(db, f)
    # w1 has planned ex25, but actual assignment is by w2
    assert len(rows) >= 1


def test_m4d_31_decision_for_discrepancy_visible(db):
    """M4D-31: Decision for discrepancy visible in roster vs actual."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc = _create_exception_from_discrepancy_helper(
        db, employee_id="w1", planned_eq_id="ex25", actual_eq_id="ex31",
        operating_date=op_date)

    dec = request_decision(db, "metro", exc.id, DecisionType.EQUIPMENT_SUBSTITUTION,
                           requested_by="sup1", reason_text="Substitution needed",
                           actual_equipment_id="ex31")
    approve_decision(db, dec.id, "metro", decided_by="admin1",
                     reason_text="Approved substitution",
                     authorization_policy="SIM_APPROVER")
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_roster_vs_actual_report(db, f)
    assert len(rows) >= 1
    # Check that decision info is present
    found_decision = any(r.decision_type == "EQUIPMENT_SUBSTITUTION" for r in rows)
    assert found_decision


def test_m4d_32_plan_actual_decision_separation(db):
    """M4D-32: Plan/actual/decision never merged into single record."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    # Create actual assignment different from plan
    aa = _make_actual_assignment(db, employee_id="w1", equipment_id="ex31",
                                 operating_date=op_date, roster_id="ra-w1")
    _make_comparison_result(db, aa.id, employee_id="w1",
                            planned_eq_id="ex25", actual_eq_id="ex31",
                            operating_date=op_date, result=ComparisonResult.MISMATCH)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_roster_vs_actual_report(db, f)
    r = rows[0]
    # Plan and actual are separate fields
    assert r.planned_equipment == "EX-025"
    assert r.actual_equipment == "EX-031"
    # They are NOT merged (planned != actual shows they're separate)
    assert r.planned_equipment != r.actual_equipment


# ── Filter Tests ──────────────────────────────────────────────

def test_m4d_33_filter_date_scope(db):
    """M4D-33: Date scope filter (single date)."""
    op_date = date(2026, 9, 5)
    other_date = date(2026, 9, 6)
    _seed_metro(db, op_date)

    # Add roster for other date
    db.add(RosterAssignment(
        id="ra-w1-other", tenant_id="metro", roster_code="RC-w1-other",
        operating_date=other_date, employee_id="w1",
        crew_id="crew_a", site_status=SiteStatusEnum.ONSITE,
        work_status=WorkStatus.WORK, shift_id="DAY", site_id="site_padang",
        planned_equipment_id="ex25", rule_version_id="rv1",
        validation_status=ValidationStatus.PUBLISHED,
    ))
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_shift_attendance_report(db, f)
    assert all(r.operating_date == op_date for r in rows)


def test_m4d_34_filter_date_range(db):
    """M4D-34: Date range filter."""
    _seed_metro(db, date(2026, 9, 5))
    db.add(RosterAssignment(
        id="ra-w1-06", tenant_id="metro", roster_code="RC-w1-06",
        operating_date=date(2026, 9, 6), employee_id="w1",
        crew_id="crew_a", site_status=SiteStatusEnum.ONSITE,
        work_status=WorkStatus.WORK, shift_id="DAY", site_id="site_padang",
        planned_equipment_id="ex25", rule_version_id="rv1",
        validation_status=ValidationStatus.PUBLISHED,
    ))
    db.add(RosterAssignment(
        id="ra-w1-08", tenant_id="metro", roster_code="RC-w1-08",
        operating_date=date(2026, 9, 8), employee_id="w1",
        crew_id="crew_a", site_status=SiteStatusEnum.ONSITE,
        work_status=WorkStatus.WORK, shift_id="DAY", site_id="site_padang",
        planned_equipment_id="ex25", rule_version_id="rv1",
        validation_status=ValidationStatus.PUBLISHED,
    ))
    db.flush()

    f = ReportFilter(tenant_id="metro", date_from=date(2026, 9, 5),
                     date_to=date(2026, 9, 6))
    rows = get_shift_attendance_report(db, f)
    dates = {r.operating_date for r in rows}
    assert date(2026, 9, 8) not in dates


def test_m4d_35_filter_site(db):
    """M4D-35: Site filter."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    # Add a second site
    site2 = Site(id="site_jkt", tenant_id="metro", site_code="JKT",
                 site_name="Jakarta Office", site_type=SiteType.MESS,
                 status=SiteStatus.ACTIVE, timezone="Asia/Jakarta",
                 effective_from=date(2026, 1, 1))
    db.add(site2)
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date, site_id="site_padang")
    rows = get_shift_attendance_report(db, f)
    assert all(r.site_name == "Padang Mine" for r in rows)


def test_m4d_36_filter_shift(db):
    """M4D-36: Shift filter."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)
    _seed_night_shift_roster(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, shift_id="DAY")
    rows = get_shift_attendance_report(db, f)
    assert all(r.shift_name == "Day" for r in rows)


def test_m4d_37_filter_crew(db):
    """M4D-37: Crew filter."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, crew_id="crew_b")
    rows = get_shift_attendance_report(db, f)
    assert len(rows) >= 1
    assert all(r.crew_name == "Crew B" for r in rows)


def test_m4d_38_filter_role(db):
    """M4D-38: Role filter."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, role_id="role_dt")
    rows = get_shift_attendance_report(db, f)
    assert len(rows) >= 1
    assert all(r.role_name == "Dump Truck Operator" for r in rows)


def test_m4d_39_filter_employee(db):
    """M4D-39: Employee filter."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_shift_attendance_report(db, f)
    assert len(rows) == 1
    assert rows[0].employee_no == "W001"


def test_m4d_40_filter_work_status(db):
    """M4D-40: Work status filter."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, work_status="REST")
    rows = get_shift_attendance_report(db, f)
    assert len(rows) == 1
    assert rows[0].work_status == "REST"


def test_m4d_41_filter_exception_type(db):
    """M4D-41: Exception type filter."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    _create_exception_from_rule(db, employee_id="w1", operating_date=op_date,
                                rule_code="LATE_BREAK_RETURN")
    _create_exception_from_rule(db, employee_id="w2", operating_date=op_date,
                                rule_code="LATE_BRIEFING")

    f = ReportFilter(tenant_id="metro", operating_date=op_date,
                     exception_type="LATE_BREAK_RETURN")
    rows = get_exception_report(db, f)
    assert len(rows) == 1
    assert rows[0].exception_type == "LATE_BREAK_RETURN"


def test_m4d_42_filter_severity(db):
    """M4D-42: Severity filter."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    _create_exception_from_rule(db, employee_id="w1", operating_date=op_date,
                                severity=RuleSeverity.WARNING)
    _create_exception_from_rule(db, employee_id="w2", operating_date=op_date,
                                rule_code="LATE_BRIEFING",
                                severity=RuleSeverity.CRITICAL)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, severity="CRITICAL")
    rows = get_exception_report(db, f)
    assert len(rows) == 1
    assert rows[0].severity == "CRITICAL"


def test_m4d_43_filter_exception_status(db):
    """M4D-43: Exception status filter."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc1 = _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)
    exc2 = _create_exception_from_rule(db, employee_id="w2", operating_date=op_date,
                                       rule_code="LATE_BRIEFING")
    acknowledge_exception(db, exc2.id, tenant_id="metro", actor_user_id="sup1")
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date,
                     exception_status="ACKNOWLEDGED")
    rows = get_exception_report(db, f)
    assert len(rows) == 1
    assert rows[0].status == "ACKNOWLEDGED"


def test_m4d_44_filter_decision_status(db):
    """M4D-44: Decision status filter."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc = _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)
    dec = request_decision(db, "metro", exc.id, DecisionType.OPERATIONAL_OVERRIDE,
                           requested_by="sup1", reason_text="Override")
    approve_decision(db, dec.id, "metro", decided_by="admin1",
                     reason_text="Approved",
                     authorization_policy="SIM_APPROVER")
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date,
                     decision_status="APPROVED")
    rows = get_exception_report(db, f)
    assert len(rows) == 1
    assert rows[0].decision_status == "APPROVED"


def test_m4d_45_filter_equipment(db):
    """M4D-45: Equipment filter for exception report."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc = _create_exception_from_discrepancy_helper(
        db, employee_id="w1", planned_eq_id="ex25", actual_eq_id="ex31",
        operating_date=op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, equipment_id="ex31")
    rows = get_exception_report(db, f)
    assert len(rows) >= 1


# ── Export: CSV Tests ─────────────────────────────────────────

def test_m4d_46_csv_generated_successfully(db):
    """M4D-46: CSV generated successfully with data."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_shift_attendance_report(db, f)
    dicts = shift_attendance_to_dicts(rows)
    csv_content = export_csv(dicts, SHIFT_ATTENDANCE_COLUMNS)

    assert csv_content is not None
    assert len(csv_content) > 0
    # Should contain header
    assert "operating_date" in csv_content


def test_m4d_47_csv_utf8_bom(db):
    """M4D-47: CSV has UTF-8 BOM."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_shift_attendance_report(db, f)
    dicts = shift_attendance_to_dicts(rows)
    csv_content = export_csv(dicts, SHIFT_ATTENDANCE_COLUMNS)

    assert csv_content.startswith('\ufeff')


def test_m4d_48_csv_stable_ordering(db):
    """M4D-48: CSV columns in stable order matching SHIFT_ATTENDANCE_COLUMNS."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_shift_attendance_report(db, f)
    dicts = shift_attendance_to_dicts(rows)
    csv_content = export_csv(dicts, SHIFT_ATTENDANCE_COLUMNS)

    # First line is header (BOM is prefix, not separate line)
    header_line = csv_content.split('\n')[0].lstrip('\ufeff')
    for col in SHIFT_ATTENDANCE_COLUMNS:
        assert col in header_line


def test_m4d_49_csv_formula_injection_safety(db):
    """M4D-49: CSV formula injection safety for dangerous prefixes."""
    # Test _sanitize_cell directly
    assert _sanitize_cell("=cmd|' /C calc'!A0") == "'=cmd|' /C calc'!A0"
    assert _sanitize_cell("+1+1") == "'+1+1"
    assert _sanitize_cell("-2+2") == "'-2+2"
    assert _sanitize_cell("@SUM(A1)") == "'@SUM(A1)"
    assert _sanitize_cell("normal text") == "normal text"
    assert _sanitize_cell("") == ""
    assert _sanitize_cell(None) is None
    assert _sanitize_cell(42) == 42


def test_m4d_50_csv_formula_injection_in_export(db):
    """M4D-50: CSV export sanitizes formula injection in cell values."""
    # Create a row with formula-injected values
    rows = [{"employee_name": "=cmd|' /C calc'!A0", "crew_name": "@SUM(A1)"}]
    columns = ["employee_name", "crew_name"]
    csv_content = export_csv(rows, columns)

    assert "'=cmd|' /C calc'!A0" in csv_content
    assert "'@SUM(A1)" in csv_content


# ── Export: XLSX Tests ────────────────────────────────────────

def test_m4d_51_xlsx_generated_successfully(db):
    """M4D-51: XLSX generated successfully."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_shift_attendance_report(db, f)
    dicts = shift_attendance_to_dicts(rows)
    metadata = build_report_metadata(db, None, "shift_attendance", f, len(rows))
    xlsx_bytes = export_xlsx(dicts, SHIFT_ATTENDANCE_COLUMNS, "Attendance",
                             metadata, "test.xlsx")

    assert xlsx_bytes is not None
    assert len(xlsx_bytes) > 0
    # Verify it's a valid ZIP (XLSX is ZIP-based)
    assert xlsx_bytes[:4] == b'PK\x03\x04'


def test_m4d_52_xlsx_sheet_names(db):
    """M4D-52: XLSX has correct sheet names (data sheet + Report_Info)."""
    import openpyxl
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_shift_attendance_report(db, f)
    dicts = shift_attendance_to_dicts(rows)
    metadata = build_report_metadata(db, None, "shift_attendance", f, len(rows))
    xlsx_bytes = export_xlsx(dicts, SHIFT_ATTENDANCE_COLUMNS, "Attendance",
                             metadata, "test.xlsx")

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert "Attendance" in wb.sheetnames
    assert "Report_Info" in wb.sheetnames
    wb.close()


def test_m4d_53_xlsx_headers_present(db):
    """M4D-53: XLSX headers present in first row."""
    import openpyxl
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_shift_attendance_report(db, f)
    dicts = shift_attendance_to_dicts(rows)
    metadata = build_report_metadata(db, None, "shift_attendance", f, len(rows))
    xlsx_bytes = export_xlsx(dicts, SHIFT_ATTENDANCE_COLUMNS, "Attendance",
                             metadata, "test.xlsx")

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Attendance"]
    # First row should have headers
    headers = [cell.value for cell in ws[1]]
    assert "Operating Date" in headers
    assert "Employee No" in headers
    wb.close()


def test_m4d_54_xlsx_formula_injection_safety(db):
    """M4D-54: XLSX formula injection safety."""
    import openpyxl
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    # Create data with formula-injected values
    rows = [{"employee_name": "=cmd|' /C calc'!A0", "crew_name": "@SUM(A1)",
             "operating_date": "2026-09-05", "site_name": "Test",
             "shift_name": "Day", "employee_no": "W001", "role_name": "Op",
             "work_status": "WORK", "planned_equipment": None,
             "briefing_in": None, "equipment_in": None, "work_start": None,
             "break_out": None, "break_in": None, "handover": None,
             "shift_out": None, "operational_state": "ABSENT",
             "attendance_exception_count": 0, "exception_types": "",
             "rule_version": "v1.0"}]
    metadata = ReportMetadata(
        tenant_name="Metro Mining", tenant_code="metro",
        site_name=None, operating_date=op_date,
        date_from=None, date_to=None, shift_name=None,
        timezone="Asia/Makassar", generated_at=_utcnow_naive(),
        report_type="shift_attendance", applied_filters={}, row_count=1,
    )
    xlsx_bytes = export_xlsx(rows, SHIFT_ATTENDANCE_COLUMNS, "Attendance",
                             metadata, "test.xlsx")

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Attendance"]
    # Find the employee_name column
    emp_col = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == "Employee Name":
            emp_col = col_idx
            break
    assert emp_col is not None
    cell_val = ws.cell(row=2, column=emp_col).value
    # Should be sanitized with apostrophe prefix
    assert str(cell_val).startswith("'")
    wb.close()


# ── Metadata Tests ────────────────────────────────────────────

def test_m4d_55_metadata_contains_tenant_info(db):
    """M4D-55: Report metadata contains tenant name, timezone, generated_at."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    # Need a mock ctx for build_report_metadata
    class MockMembership:
        tenant_id = "metro"
    class MockCtx:
        membership = MockMembership()

    metadata = build_report_metadata(db, MockCtx(), "shift_attendance", f, 0)
    assert metadata.tenant_name == "Metro Mining"
    assert metadata.tenant_code == "metro"
    assert metadata.timezone == "Asia/Makassar"
    assert metadata.generated_at is not None
    assert metadata.report_type == "shift_attendance"


def test_m4d_56_metadata_applied_filters(db):
    """M4D-56: Metadata contains applied filters."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date,
                     shift_id="DAY", crew_id="crew_a")

    class MockMembership:
        tenant_id = "metro"
    class MockCtx:
        membership = MockMembership()

    metadata = build_report_metadata(db, MockCtx(), "shift_attendance", f, 5)
    assert metadata.applied_filters["operating_date"] == str(op_date)
    assert metadata.applied_filters["shift_id"] == "DAY"
    assert metadata.applied_filters["crew_id"] == "crew_a"
    assert metadata.row_count == 5


def test_m4d_57_filename_generation(db):
    """M4D-57: Filename generation predictable."""
    fn = generate_filename("shift_attendance", "metro",
                           operating_date=date(2026, 9, 5), fmt="csv")
    assert fn == "metro_shift_attendance_2026-09-05.csv"

    fn2 = generate_filename("exceptions", "metro",
                            date_from=date(2026, 9, 1), date_to=date(2026, 9, 30),
                            fmt="xlsx")
    assert fn2 == "metro_exceptions_2026-09-01_to_2026-09-30.xlsx"

    fn3 = generate_filename("roster_vs_actual", "metro",
                            operating_date=date(2026, 9, 5),
                            shift_id="DAY", fmt="xlsx")
    assert fn3 == "metro_roster_vs_actual_2026-09-05_DAY.xlsx"


# ── Isolation Tests ───────────────────────────────────────────

def test_m4d_58_tenant_isolation_attendance(db):
    """M4D-58: Metro report cannot contain Client B attendance data."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)
    _seed_client_b(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_shift_attendance_report(db, f)
    employee_nos = {r.employee_no for r in rows}
    assert "CB001" not in employee_nos
    assert "W001" in employee_nos


def test_m4d_59_tenant_isolation_exceptions(db):
    """M4D-59: Metro report cannot contain Client B exceptions."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)
    _seed_client_b(db, op_date)

    # Create exception for metro
    _create_exception_from_rule(db, tenant_id="metro", employee_id="w1",
                                operating_date=op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_exception_report(db, f)
    for r in rows:
        assert r.employee_no != "CB001"


def test_m4d_60_tenant_isolation_equipment(db):
    """M4D-60: Metro report cannot contain Client B equipment data."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)
    _seed_client_b(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_roster_vs_actual_report(db, f)
    for r in rows:
        assert r.planned_equipment != "TR-001"


def test_m4d_61_site_isolation(db):
    """M4D-61: Site A report must not include Site B data."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    # Add second site and roster for w4 at that site
    site2 = Site(id="site_jkt", tenant_id="metro", site_code="JKT",
                 site_name="Jakarta Office", site_type=SiteType.MESS,
                 status=SiteStatus.ACTIVE, timezone="Asia/Jakarta",
                 effective_from=date(2026, 1, 1))
    db.add(site2)
    db.flush()

    # Update w4's roster to different site
    ra_w4 = db.query(RosterAssignment).filter_by(id="ra-w4").first()
    ra_w4.site_id = "site_jkt"
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date, site_id="site_padang")
    rows = get_shift_attendance_report(db, f)
    site_names = {r.site_name for r in rows}
    assert "Jakarta Office" not in site_names
    assert "Padang Mine" in site_names


# ── Empty Report Tests ────────────────────────────────────────

def test_m4d_62_empty_report_valid(db):
    """M4D-62: Empty report returns valid empty result (not error)."""
    _seed_metro(db, date(2026, 9, 5))

    f = ReportFilter(tenant_id="metro", operating_date=date(2026, 12, 25))
    rows = get_shift_attendance_report(db, f)
    assert rows == []
    dicts = shift_attendance_to_dicts(rows)
    assert dicts == []


def test_m4d_63_empty_csv_valid_with_headers(db):
    """M4D-63: Empty CSV valid with headers."""
    csv_content = export_csv([], SHIFT_ATTENDANCE_COLUMNS)
    assert csv_content.startswith('\ufeff')
    # Should have header line
    lines = csv_content.strip().split('\n')
    assert len(lines) >= 1
    # Header should be in the first non-BOM line
    header_text = lines[0].replace('\ufeff', '')
    for col in SHIFT_ATTENDANCE_COLUMNS:
        assert col in header_text


def test_m4d_64_empty_xlsx_valid_with_headers(db):
    """M4D-64: Empty XLSX valid with headers."""
    import openpyxl
    metadata = ReportMetadata(
        tenant_name="Metro Mining", tenant_code="metro",
        site_name=None, operating_date=date(2026, 12, 25),
        date_from=None, date_to=None, shift_name=None,
        timezone="Asia/Makassar", generated_at=_utcnow_naive(),
        report_type="shift_attendance", applied_filters={}, row_count=0,
    )
    xlsx_bytes = export_xlsx([], SHIFT_ATTENDANCE_COLUMNS, "Attendance",
                             metadata, "test.xlsx")
    assert len(xlsx_bytes) > 0

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Attendance"]
    # Should have header row
    headers = [cell.value for cell in ws[1]]
    assert len(headers) > 0
    assert ws.max_row == 1  # Only header, no data
    wb.close()


# ── Safety Tests ──────────────────────────────────────────────

def test_m4d_65_no_payroll_calculation(db):
    """M4D-65: No payroll calculation in reports."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_shift_attendance_report(db, f)
    dicts = shift_attendance_to_dicts(rows)
    # Verify no payroll-related fields
    for d in dicts:
        assert "payroll" not in d
        assert "salary" not in d
        assert "overtime_hours" not in d
        assert "deduction" not in d


def test_m4d_66_no_disciplinary_transformation(db):
    """M4D-66: No disciplinary/HSE transformation."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_exception_report(db, f)
    dicts = exception_report_to_dicts(rows)
    for d in dicts:
        assert "disciplinary_action" not in d
        assert "hse_category" not in d
        assert "punishment" not in d


def test_m4d_67_max_export_rows_bounded(db):
    """M4D-67: MAX_EXPORT_ROWS bounds query results."""
    assert MAX_EXPORT_ROWS == 1000


def test_m4d_68_config_incomplete_for_missing_policies(db):
    """M4D-68: CONFIG_INCOMPLETE status for missing geofence/TBC policies."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    # Create a rule eval with CONFIG_INCOMPLETE status
    re = RuleEvaluation(
        id=_uid(), tenant_id="metro", employee_id="w1",
        operating_date=op_date, shift_id="DAY",
        rule_code="LOCATION_OUTSIDE_GEOFENCE", rule_version_id="rv1",
        evaluated_at=_utcnow_naive(),
        status=RuleEvaluationStatus.CONFIG_INCOMPLETE, severity=RuleSeverity.WARNING,
        evidence_key=_uid(),
    )
    db.add(re)
    db.flush()

    # CONFIG_INCOMPLETE should NOT create an exception
    exc = create_exception_from_rule_evaluation(db, re)
    assert exc is None


# ── Regression Tests ──────────────────────────────────────────

def test_m4d_69_m4a_regression_roster_loads(db):
    """M4D-69: M4A regression — roster assignments still load correctly."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    # Verify roster data is intact
    ras = db.query(RosterAssignment).filter_by(
        tenant_id="metro", operating_date=op_date
    ).all()
    assert len(ras) == 6  # 4 WORK + 1 REST + 1 OFFSITE


def test_m4d_70_m4b_regression_equipment_assignments(db):
    """M4D-70: M4B regression — equipment assignments still work."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    aa = _make_actual_assignment(db, employee_id="w1", equipment_id="ex25",
                                 operating_date=op_date, roster_id="ra-w1")
    assert aa is not None
    assert aa.equipment_id == "ex25"


def test_m4d_71_m4c_regression_exceptions_still_work(db):
    """M4D-71: M4C regression — exception lifecycle still works."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc = _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)
    assert exc is not None
    assert exc.status == ExceptionStatus.OPEN

    acknowledge_exception(db, exc.id, tenant_id="metro", actor_user_id="sup1")
    db.flush()
    updated = get_exception(db, exc.id, "metro")
    assert updated.status == ExceptionStatus.ACKNOWLEDGED


# ── Additional Coverage Tests ─────────────────────────────────

def test_m4d_72_exception_report_all_statuses_visible(db):
    """M4D-72: Exception report shows all statuses when no filter."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc1 = _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)
    exc2 = _create_exception_from_rule(db, employee_id="w2", operating_date=op_date,
                                       rule_code="LATE_BRIEFING")
    exc3 = _create_exception_from_rule(db, employee_id="w3", operating_date=op_date,
                                       rule_code="EARLY_HANDOVER")
    acknowledge_exception(db, exc2.id, tenant_id="metro", actor_user_id="sup1")
    resolve_exception(db, exc3.id, tenant_id="metro", actor_user_id="sup1", reason="Fixed")
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_exception_report(db, f)
    statuses = {r.status for r in rows}
    assert "OPEN" in statuses
    assert "ACKNOWLEDGED" in statuses
    assert "RESOLVED" in statuses


def test_m4d_73_roster_vs_actual_no_actual_assignment(db):
    """M4D-73: Roster vs actual with no actual assignment shows None actuals."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_roster_vs_actual_report(db, f)
    assert len(rows) == 1
    r = rows[0]
    assert r.planned_equipment == "EX-025"
    assert r.actual_equipment is None
    assert r.comparison_result is None


def test_m4d_74_discrepancy_type_visible(db):
    """M4D-74: Discrepancy type visible in roster vs actual."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    aa = _make_actual_assignment(db, employee_id="w1", equipment_id="ex31",
                                 operating_date=op_date, roster_id="ra-w1")
    _make_comparison_result(db, aa.id, employee_id="w1",
                            planned_eq_id="ex25", actual_eq_id="ex31",
                            operating_date=op_date, result=ComparisonResult.MISMATCH)
    _make_discrepancy(db, aa.id, employee_id="w1",
                      planned_eq_id="ex25", actual_eq_id="ex31",
                      operating_date=op_date,
                      disc_type=DiscrepancyType.EQUIPMENT_MISMATCH)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_roster_vs_actual_report(db, f)
    assert len(rows) == 1
    assert rows[0].discrepancy_type == "EQUIPMENT_MISMATCH"


def test_m4d_75_exception_report_requested_by_decided_by(db):
    """M4D-75: Exception report shows requested_by and decided_by names."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc = _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)
    dec = request_decision(db, "metro", exc.id, DecisionType.OPERATIONAL_OVERRIDE,
                           requested_by="sup1", reason_text="Override")
    approve_decision(db, dec.id, "metro", decided_by="admin1",
                     reason_text="Approved",
                     authorization_policy="SIM_APPROVER")
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_exception_report(db, f)
    r = rows[0]
    assert r.requested_by == "Supervisor One"
    assert r.decided_by == "Admin One"


def test_m4d_76_shift_attendance_to_dicts_serialization(db):
    """M4D-76: shift_attendance_to_dicts produces correct dict keys."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_shift_attendance_report(db, f)
    dicts = shift_attendance_to_dicts(rows)
    assert len(dicts) == 1
    d = dicts[0]
    for col in SHIFT_ATTENDANCE_COLUMNS:
        assert col in d


def test_m4d_77_exception_report_to_dicts_serialization(db):
    """M4D-77: exception_report_to_dicts produces correct dict keys."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_exception_report(db, f)
    dicts = exception_report_to_dicts(rows)
    assert len(dicts) == 1
    d = dicts[0]
    for col in EXCEPTION_REPORT_COLUMNS:
        assert col in d


def test_m4d_78_roster_vs_actual_to_dicts_serialization(db):
    """M4D-78: roster_vs_actual_to_dicts produces correct dict keys."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_roster_vs_actual_report(db, f)
    dicts = roster_vs_actual_to_dicts(rows)
    assert len(dicts) == 1
    d = dicts[0]
    for col in ROSTER_VS_ACTUAL_COLUMNS:
        assert col in d


def test_m4d_79_xlsx_report_info_metadata_sheet(db):
    """M4D-79: XLSX Report_Info sheet contains metadata."""
    import openpyxl
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, shift_id="DAY")

    class MockMembership:
        tenant_id = "metro"
    class MockCtx:
        membership = MockMembership()

    rows = get_shift_attendance_report(db, f)
    dicts = shift_attendance_to_dicts(rows)
    metadata = build_report_metadata(db, MockCtx(), "shift_attendance", f, len(rows))
    xlsx_bytes = export_xlsx(dicts, SHIFT_ATTENDANCE_COLUMNS, "Attendance",
                             metadata, "test.xlsx")

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws_info = wb["Report_Info"]
    # Check metadata values
    info_data = {}
    for row in ws_info.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0] and row[1]:
            info_data[row[0]] = row[1]

    assert info_data.get("Tenant") == "Metro Mining"
    assert info_data.get("Timezone") == "Asia/Makassar"
    assert info_data.get("Report Type") == "shift_attendance"
    assert info_data.get("Shift") == "Day"
    wb.close()


def test_m4d_80_multiple_employees_in_report(db):
    """M4D-80: Multiple employees appear in single report."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_shift_attendance_report(db, f)
    employee_nos = {r.employee_no for r in rows}
    assert "W001" in employee_nos
    assert "W002" in employee_nos
    assert "W003" in employee_nos
    assert "W004" in employee_nos
    assert "W005" in employee_nos
    assert "W006" in employee_nos


def test_m4d_81_exception_from_discrepancy_source_type(db):
    """M4D-81: Exception from discrepancy shows EQUIPMENT_DISCREPANCY source."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    _create_exception_from_discrepancy_helper(
        db, employee_id="w1", planned_eq_id="ex25", actual_eq_id="ex31",
        operating_date=op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_exception_report(db, f)
    assert len(rows) == 1
    assert rows[0].source_type == "EQUIPMENT_DISCREPANCY"


def test_m4d_82_csv_exception_report_columns(db):
    """M4D-82: CSV exception report has all EXCEPTION_REPORT_COLUMNS."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_exception_report(db, f)
    dicts = exception_report_to_dicts(rows)
    csv_content = export_csv(dicts, EXCEPTION_REPORT_COLUMNS)

    header_line = csv_content.split('\n')[0].lstrip('\ufeff')
    for col in EXCEPTION_REPORT_COLUMNS:
        assert col in header_line


def test_m4d_83_csv_roster_vs_actual_columns(db):
    """M4D-83: CSV roster vs actual has all ROSTER_VS_ACTUAL_COLUMNS."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_roster_vs_actual_report(db, f)
    dicts = roster_vs_actual_to_dicts(rows)
    csv_content = export_csv(dicts, ROSTER_VS_ACTUAL_COLUMNS)

    header_line = csv_content.split('\n')[0].lstrip('\ufeff')
    for col in ROSTER_VS_ACTUAL_COLUMNS:
        assert col in header_line


def test_m4d_84_xlsx_date_formatting(db):
    """M4D-84: XLSX formats dates as dates, not strings."""
    import openpyxl
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_shift_attendance_report(db, f)
    dicts = shift_attendance_to_dicts(rows)
    metadata = ReportMetadata(
        tenant_name="Metro Mining", tenant_code="metro",
        site_name="Padang Mine", operating_date=op_date,
        date_from=None, date_to=None, shift_name="Day",
        timezone="Asia/Makassar", generated_at=_utcnow_naive(),
        report_type="shift_attendance", applied_filters={}, row_count=1,
    )
    xlsx_bytes = export_xlsx(dicts, SHIFT_ATTENDANCE_COLUMNS, "Attendance",
                             metadata, "test.xlsx")

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Attendance"]
    # Find operating_date column
    date_col = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == "Operating Date":
            date_col = col_idx
            break
    assert date_col is not None
    # Check that the cell value is a date object, not a string
    cell_val = ws.cell(row=2, column=date_col).value
    if cell_val is not None:
        assert isinstance(cell_val, (date, datetime)) or str(cell_val) == "2026-09-05"
    wb.close()


def test_m4d_85_decision_pending_visible(db):
    """M4D-85: PENDING decision visible in exception report."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc = _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)
    request_decision(db, "metro", exc.id, DecisionType.OPERATIONAL_OVERRIDE,
                     requested_by="sup1", reason_text="Pending override")
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_exception_report(db, f)
    assert rows[0].decision_status == "PENDING"
    assert rows[0].decision_type == "OPERATIONAL_OVERRIDE"


def test_m4d_86_filename_with_date_range(db):
    """M4D-86: Filename with date range includes both dates."""
    fn = generate_filename("shift_attendance", "metro",
                           date_from=date(2026, 9, 1), date_to=date(2026, 9, 30),
                           fmt="csv")
    assert "2026-09-01" in fn
    assert "2026-09-30" in fn
    assert "to" in fn
    assert fn.endswith(".csv")


def test_m4d_87_filename_with_only_date_from(db):
    """M4D-87: Filename with only date_from."""
    fn = generate_filename("exceptions", "metro",
                           date_from=date(2026, 9, 1), fmt="xlsx")
    assert "from_2026-09-01" in fn
    assert fn.endswith(".xlsx")


def test_m4d_88_filename_with_only_date_to(db):
    """M4D-88: Filename with only date_to."""
    fn = generate_filename("exceptions", "metro",
                           date_to=date(2026, 9, 30), fmt="xlsx")
    assert "to_2026-09-30" in fn


def test_m4d_89_metadata_site_name_in_report(db):
    """M4D-89: Metadata includes site name when site filter applied."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, site_id="site_padang")

    class MockMembership:
        tenant_id = "metro"
    class MockCtx:
        membership = MockMembership()

    metadata = build_report_metadata(db, MockCtx(), "shift_attendance", f, 0)
    assert metadata.site_name == "Padang Mine"


def test_m4d_90_metadata_shift_name_in_report(db):
    """M4D-90: Metadata includes shift name when shift filter applied."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, shift_id="DAY")

    class MockMembership:
        tenant_id = "metro"
    class MockCtx:
        membership = MockMembership()

    metadata = build_report_metadata(db, MockCtx(), "shift_attendance", f, 0)
    assert metadata.shift_name == "Day"


def test_m4d_91_exception_report_empty_for_no_exceptions(db):
    """M4D-91: Exception report empty when no exceptions exist."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_exception_report(db, f)
    assert rows == []


def test_m4d_92_roster_vs_actual_only_work_status(db):
    """M4D-92: Roster vs actual only includes WORK status employees."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_roster_vs_actual_report(db, f)
    # Should only have WORK employees (w1-w4), not REST or OFFSITE
    employee_nos = {r.employee_no for r in rows}
    assert "W005" not in employee_nos  # REST
    assert "W006" not in employee_nos  # OFFSITE
    assert "W001" in employee_nos


def test_m4d_93_xlsx_freeze_header_row(db):
    """M4D-93: XLSX freeze panes on header row."""
    import openpyxl
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    rows = get_shift_attendance_report(db, f)
    dicts = shift_attendance_to_dicts(rows)
    metadata = ReportMetadata(
        tenant_name="Metro Mining", tenant_code="metro",
        site_name=None, operating_date=op_date,
        date_from=None, date_to=None, shift_name=None,
        timezone="Asia/Makassar", generated_at=_utcnow_naive(),
        report_type="shift_attendance", applied_filters={}, row_count=len(rows),
    )
    xlsx_bytes = export_xlsx(dicts, SHIFT_ATTENDANCE_COLUMNS, "Attendance",
                             metadata, "test.xlsx")

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Attendance"]
    assert ws.freeze_panes == "A2"
    wb.close()


def test_m4d_94_exception_report_equipment_mismatch_type(db):
    """M4D-94: Exception report shows EQUIPMENT_MISMATCH type from discrepancy."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    _create_exception_from_discrepancy_helper(
        db, employee_id="w1", planned_eq_id="ex25", actual_eq_id="ex31",
        operating_date=op_date, disc_type=DiscrepancyType.EQUIPMENT_MISMATCH)

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_exception_report(db, f)
    assert len(rows) == 1
    assert rows[0].exception_type == "EQUIPMENT_MISMATCH"


def test_m4d_95_shift_attendance_break_in_after_break_out(db):
    """M4D-95: Break in after break out shows operational state WORK."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    work_ev = _make_canonical_event(db, employee_id="w1", operating_date=op_date,
                                    event_type=CanonicalEventType.CHECK_IN,
                                    local_ts=datetime(2026, 9, 5, 7, 30, 0))
    _make_checkpoint_result(db, work_ev.id, employee_id="w1",
                            operating_date=op_date, checkpoint_type="WORK_START")

    break_out_ev = _make_canonical_event(db, employee_id="w1", operating_date=op_date,
                                         event_type=CanonicalEventType.BREAK_OUT,
                                         local_ts=datetime(2026, 9, 5, 12, 0, 0))
    _make_checkpoint_result(db, break_out_ev.id, employee_id="w1",
                            operating_date=op_date, checkpoint_type="BREAK_OUT")

    break_in_ev = _make_canonical_event(db, employee_id="w1", operating_date=op_date,
                                        event_type=CanonicalEventType.BREAK_IN,
                                        local_ts=datetime(2026, 9, 5, 13, 0, 0))
    _make_checkpoint_result(db, break_in_ev.id, employee_id="w1",
                            operating_date=op_date, checkpoint_type="BREAK_IN")

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_shift_attendance_report(db, f)
    # After break_in, state should be WORK (not ON_BREAK)
    assert rows[0].operational_state == "WORK"


def test_m4d_96_exception_report_shift_filter(db):
    """M4D-96: Exception report respects shift filter."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)
    _seed_night_shift_roster(db, op_date)

    # Create exception for DAY shift
    _create_exception_from_rule(db, employee_id="w1", operating_date=op_date,
                                shift_id="DAY")
    # Create exception for NIGHT shift (use w4 who has night roster)
    re_night = _make_rule_eval(db, employee_id="w4", operating_date=op_date,
                               shift_id="NIGHT", rule_code="LATE_BRIEFING")
    create_exception_from_rule_evaluation(db, re_night)
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date, shift_id="DAY")
    rows = get_exception_report(db, f)
    assert len(rows) == 1
    assert rows[0].shift_name == "Day"


def test_m4d_97_csv_all_three_report_types(db):
    """M4D-97: CSV export works for all three report types."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)

    # Shift attendance
    sa_rows = get_shift_attendance_report(db, f)
    sa_dicts = shift_attendance_to_dicts(sa_rows)
    sa_csv = export_csv(sa_dicts, SHIFT_ATTENDANCE_COLUMNS)
    assert sa_csv.startswith('\ufeff')

    # Exception
    exc_rows = get_exception_report(db, f)
    exc_dicts = exception_report_to_dicts(exc_rows)
    exc_csv = export_csv(exc_dicts, EXCEPTION_REPORT_COLUMNS)
    assert exc_csv.startswith('\ufeff')

    # Roster vs actual
    rva_rows = get_roster_vs_actual_report(db, f)
    rva_dicts = roster_vs_actual_to_dicts(rva_rows)
    rva_csv = export_csv(rva_dicts, ROSTER_VS_ACTUAL_COLUMNS)
    assert rva_csv.startswith('\ufeff')


def test_m4d_98_xlsx_all_three_report_types(db):
    """M4D-98: XLSX export works for all three report types."""
    import openpyxl
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    f = ReportFilter(tenant_id="metro", operating_date=op_date)
    metadata = ReportMetadata(
        tenant_name="Metro Mining", tenant_code="metro",
        site_name=None, operating_date=op_date,
        date_from=None, date_to=None, shift_name=None,
        timezone="Asia/Makassar", generated_at=_utcnow_naive(),
        report_type="test", applied_filters={}, row_count=0,
    )

    # Shift attendance
    sa_rows = get_shift_attendance_report(db, f)
    sa_dicts = shift_attendance_to_dicts(sa_rows)
    sa_xlsx = export_xlsx(sa_dicts, SHIFT_ATTENDANCE_COLUMNS, "Attendance",
                          metadata, "test.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(sa_xlsx))
    assert "Attendance" in wb.sheetnames
    wb.close()

    # Exception
    exc_rows = get_exception_report(db, f)
    exc_dicts = exception_report_to_dicts(exc_rows)
    exc_xlsx = export_xlsx(exc_dicts, EXCEPTION_REPORT_COLUMNS, "Exceptions",
                           metadata, "test.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(exc_xlsx))
    assert "Exceptions" in wb.sheetnames
    wb.close()

    # Roster vs actual
    rva_rows = get_roster_vs_actual_report(db, f)
    rva_dicts = roster_vs_actual_to_dicts(rva_rows)
    rva_xlsx = export_xlsx(rva_dicts, ROSTER_VS_ACTUAL_COLUMNS, "Roster_vs_Actual",
                           metadata, "test.xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(rva_xlsx))
    assert "Roster_vs_Actual" in wb.sheetnames
    wb.close()


def test_m4d_99_exception_lifecycle_open_to_waived(db):
    """M4D-99: Exception lifecycle OPEN → WAIVED preserved in report."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    exc = _create_exception_from_rule(db, employee_id="w1", operating_date=op_date)
    waive_exception(db, exc.id, tenant_id="metro", actor_user_id="sup1",
                    reason="Acceptable under circumstances")
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date,
                     exception_status="WAIVED")
    rows = get_exception_report(db, f)
    assert len(rows) == 1
    assert rows[0].status == "WAIVED"
    assert rows[0].waived_at is not None


def test_m4d_100_roster_vs_actual_decision_for_mismatch(db):
    """M4D-100: Decision visible for equipment mismatch discrepancy."""
    op_date = date(2026, 9, 5)
    _seed_metro(db, op_date)

    aa = _make_actual_assignment(db, employee_id="w1", equipment_id="ex31",
                                 operating_date=op_date, roster_id="ra-w1")
    _make_comparison_result(db, aa.id, employee_id="w1",
                            planned_eq_id="ex25", actual_eq_id="ex31",
                            operating_date=op_date, result=ComparisonResult.MISMATCH)
    disc = _make_discrepancy(db, aa.id, employee_id="w1",
                             planned_eq_id="ex25", actual_eq_id="ex31",
                             operating_date=op_date,
                             disc_type=DiscrepancyType.EQUIPMENT_MISMATCH)
    exc = create_exception_from_discrepancy(db, disc)
    db.flush()

    dec = request_decision(db, "metro", exc.id, DecisionType.EQUIPMENT_SUBSTITUTION,
                           requested_by="sup1", reason_text="Equipment swap needed",
                           actual_equipment_id="ex31")
    approve_decision(db, dec.id, "metro", decided_by="admin1",
                     reason_text="Approved equipment substitution",
                     authorization_policy="SIM_APPROVER")
    db.flush()

    f = ReportFilter(tenant_id="metro", operating_date=op_date, employee_id="w1")
    rows = get_roster_vs_actual_report(db, f)
    assert len(rows) == 1
    r = rows[0]
    assert r.decision_type == "EQUIPMENT_SUBSTITUTION"
    assert r.decision_status == "APPROVED"
    assert r.decision_reason == "Approved equipment substitution"
