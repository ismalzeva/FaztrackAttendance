import json
"""Admin Management API — Fase 1 HRD Lumin.

Endpoints:
- CRUD karyawan (create/update/deactivate)
- Assign karyawan ke proyek
- Reset PIN
- Review approval (enhanced)
- Laporan kehadiran + export
"""
import io
from datetime import date, datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import func, select, update
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import RequestContext, admin_context, approval_context, tenant_context
from app.models import (
    Assignment, AttendanceEvent, AttendanceStatus, AttendanceType,
    AuditEvent, Project, RoleCode, Tenant, User, WorkSchedule, Worker,
)
from app.security import hash_password

router = APIRouter(prefix="/api/v1/admin", tags=["admin-management"])


def _envelope(data, request: Request):
    return {
        "data": data,
        "meta": {
            "correlation_id": request.state.correlation_id,
            "server_time": datetime.now(timezone.utc).isoformat(),
            "version": "v1",
        },
    }


# ──────────────────────────────────────────────
# KARYAWAN CRUD
# ──────────────────────────────────────────────

class WorkerCreate(BaseModel):
    code: str = Field(min_length=1, max_length=50)
    name: str = Field(min_length=1, max_length=200)
    phone: str | None = None
    pin: str = Field(min_length=4, max_length=32)


class WorkerUpdate(BaseModel):
    name: str | None = None
    phone: str | None = None
    is_active: bool | None = None


class PinReset(BaseModel):
    new_pin: str = Field(min_length=4, max_length=32)


@router.get("/workers")
def list_workers(
    request: Request,
    include_inactive: bool = Query(False),
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """List semua karyawan (opsional termasuk yang nonaktif)."""
    q = select(Worker).where(Worker.tenant_id == ctx.membership.tenant_id)
    if not include_inactive:
        q = q.where(Worker.is_active.is_(True))
    q = q.order_by(Worker.code)
    workers = db.scalars(q).all()
    return _envelope(
        [
            {
                "id": w.id,
                "code": w.code,
                "name": w.name,
                "phone": w.phone,
                "is_active": w.is_active,
                "has_pin": bool(w.pin_hash),
            }
            for w in workers
        ],
        request,
    )


@router.post("/workers", status_code=201)
def create_worker(
    body: WorkerCreate,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Tambah karyawan baru."""
    tid = ctx.membership.tenant_id
    existing = db.scalar(
        select(Worker).where(Worker.tenant_id == tid, Worker.code == body.code.upper())
    )
    if existing:
        raise HTTPException(409, detail={"code": "WORKER_CODE_EXISTS"})
    w = Worker(
        tenant_id=tid,
        code=body.code.upper(),
        name=body.name,
        phone=body.phone,
        pin_hash=hash_password(body.pin),
        is_active=True,
    )
    db.add(w)
    db.commit()
    return _envelope({"id": w.id, "code": w.code, "name": w.name}, request)


@router.patch("/workers/{worker_id}")
def update_worker(
    worker_id: str,
    body: WorkerUpdate,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Edit data karyawan (nama, phone, aktif/nonaktif)."""
    w = db.scalar(
        select(Worker).where(
            Worker.id == worker_id, Worker.tenant_id == ctx.membership.tenant_id
        )
    )
    if not w:
        raise HTTPException(404, detail={"code": "WORKER_NOT_FOUND"})
    if body.name is not None:
        w.name = body.name
    if body.phone is not None:
        w.phone = body.phone
    if body.is_active is not None:
        w.is_active = body.is_active
    db.commit()
    return _envelope({"id": w.id, "code": w.code, "name": w.name, "is_active": w.is_active}, request)


@router.post("/workers/{worker_id}/reset-pin")
def reset_pin(
    worker_id: str,
    body: PinReset,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Reset PIN karyawan."""
    w = db.scalar(
        select(Worker).where(
            Worker.id == worker_id, Worker.tenant_id == ctx.membership.tenant_id
        )
    )
    if not w:
        raise HTTPException(404, detail={"code": "WORKER_NOT_FOUND"})
    w.pin_hash = hash_password(body.new_pin)
    db.commit()
    return _envelope({"id": w.id, "code": w.code, "name": w.name, "pin_reset": True}, request)


# ──────────────────────────────────────────────
# ASSIGN KARYAWAN KE PROYEK
# ──────────────────────────────────────────────

class AssignmentCreate(BaseModel):
    worker_id: str
    project_id: str
    work_date: date


@router.get("/assignments")
def list_assignments(
    request: Request,
    work_date: date = Query(...),
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """List assignment karyawan ke proyek untuk tanggal tertentu."""
    tid = ctx.membership.tenant_id
    rows = db.scalars(
        select(Assignment).where(
            Assignment.tenant_id == tid, Assignment.work_date == work_date
        )
    ).all()
    result = []
    for a in rows:
        w = db.get(Worker, a.worker_id)
        p = db.get(Project, a.project_id)
        result.append(
            {
                "assignment_id": a.id,
                "worker_id": a.worker_id,
                "worker_code": w.code if w else "?",
                "worker_name": w.name if w else "?",
                "project_id": a.project_id,
                "project_code": p.code if p else "?",
                "project_name": p.name if p else "?",
                "work_date": str(a.work_date),
            }
        )
    return _envelope(result, request)


@router.post("/assignments", status_code=201)
def create_assignment(
    body: AssignmentCreate,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Assign karyawan ke proyek untuk tanggal tertentu."""
    tid = ctx.membership.tenant_id
    # Check worker exists
    w = db.scalar(select(Worker).where(Worker.id == body.worker_id, Worker.tenant_id == tid))
    if not w:
        raise HTTPException(404, detail={"code": "WORKER_NOT_FOUND"})
    # Check project exists
    p = db.scalar(select(Project).where(Project.id == body.project_id, Project.tenant_id == tid))
    if not p:
        raise HTTPException(404, detail={"code": "PROJECT_NOT_FOUND"})
    # Check duplicate
    existing = db.scalar(
        select(Assignment).where(
            Assignment.tenant_id == tid,
            Assignment.worker_id == body.worker_id,
            Assignment.work_date == body.work_date,
        )
    )
    if existing:
        # Update project
        existing.project_id = body.project_id
        db.commit()
        return _envelope({"id": existing.id, "updated": True}, request)
    a = Assignment(
        tenant_id=tid,
        worker_id=body.worker_id,
        project_id=body.project_id,
        work_date=body.work_date,
    )
    db.add(a)
    db.commit()
    return _envelope({"id": a.id}, request)


@router.delete("/assignments/{assignment_id}")
def delete_assignment(
    assignment_id: str,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Hapus assignment."""
    a = db.scalar(
        select(Assignment).where(
            Assignment.id == assignment_id, Assignment.tenant_id == ctx.membership.tenant_id
        )
    )
    if not a:
        raise HTTPException(404, detail={"code": "ASSIGNMENT_NOT_FOUND"})
    db.delete(a)
    db.commit()
    return _envelope({"deleted": True}, request)


# ──────────────────────────────────────────────
# REVIEW APPROVAL (enhanced)
# ──────────────────────────────────────────────

class ReviewDecision(BaseModel):
    decision: str = Field(pattern="^(APPROVED|REJECTED)$")
    reason: str = Field(min_length=1, max_length=500)


@router.get("/review-queue")
def review_queue(
    request: Request,
    ctx: RequestContext = Depends(approval_context),
    db: Session = Depends(get_db),
):
    """List absen yang perlu review (status REVIEW)."""
    tid = ctx.membership.tenant_id
    events = db.scalars(
        select(AttendanceEvent).where(
            AttendanceEvent.tenant_id == tid,
            AttendanceEvent.status == AttendanceStatus.REVIEW,
        ).order_by(AttendanceEvent.server_time.desc())
    ).all()
    result = []
    for ev in events:
        w = db.get(Worker, ev.worker_id)
        p = db.get(Project, ev.project_id)
        result.append({
            "event_id": ev.id,
            "worker_code": w.code if w else "?",
            "worker_name": w.name if w else "?",
            "project_code": p.code if p else "?",
            "project_name": p.name if p else "?",
            "event_type": ev.event_type.value if hasattr(ev.event_type, 'value') else str(ev.event_type),
            "server_time": ev.server_time.isoformat(),
            "latitude": ev.latitude,
            "longitude": ev.longitude,
            "reason_code": ev.reason_code,
            "site_note": ev.site_note,
        })
    return _envelope(result, request)


@router.post("/review-queue/{event_id}/decide")
def decide_review(
    event_id: str,
    body: ReviewDecision,
    request: Request,
    ctx: RequestContext = Depends(approval_context),
    db: Session = Depends(get_db),
):
    """Approve atau reject absen REVIEW."""
    ev = db.scalar(
        select(AttendanceEvent).where(
            AttendanceEvent.id == event_id,
            AttendanceEvent.tenant_id == ctx.membership.tenant_id,
        )
    )
    if not ev:
        raise HTTPException(404, detail={"code": "EVENT_NOT_FOUND"})
    if ev.status != AttendanceStatus.REVIEW:
        raise HTTPException(409, detail={"code": "NOT_IN_REVIEW_STATUS"})
    ev.status = AttendanceStatus(body.decision)
    ev.reviewed_at = datetime.now(timezone.utc)
    ev.reviewed_by = ctx.user.id
    ev.review_reason = body.reason
    db.commit()
    return _envelope({
        "event_id": ev.id,
        "new_status": ev.status.value,
        "reviewed_by": ctx.user.display_name,
        "reviewed_at": ev.reviewed_at.isoformat(),
        "reason": ev.review_reason,
    }, request)


# ──────────────────────────────────────────────
# LAPORAN KEHADIRAN
# ──────────────────────────────────────────────

@router.get("/reports/attendance")
def attendance_report(
    request: Request,
    date_from: date = Query(...),
    date_to: date = Query(...),
    project_id: str | None = Query(None),
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Laporan kehadiran per periode."""
    tid = ctx.membership.tenant_id
    q = (
        select(AttendanceEvent)
        .where(
            AttendanceEvent.tenant_id == tid,
            AttendanceEvent.work_date >= date_from,
            AttendanceEvent.work_date <= date_to,
            AttendanceEvent.status.in_([AttendanceStatus.VALID, AttendanceStatus.REVIEW]),
        )
        .order_by(AttendanceEvent.work_date, AttendanceEvent.server_time)
    )
    if project_id:
        q = q.where(AttendanceEvent.project_id == project_id)
    events = db.scalars(q).all()

    # Group by worker + date
    from collections import defaultdict
    records = defaultdict(lambda: {"check_in": None, "check_out": None})
    for ev in events:
        key = (ev.worker_id, str(ev.work_date))
        etype = ev.event_type.value if hasattr(ev.event_type, 'value') else str(ev.event_type)
        if etype == "CHECK_IN":
            records[key]["check_in"] = ev
        elif etype == "CHECK_OUT":
            records[key]["check_out"] = ev

    # Build report rows
    rows = []
    for (wid, wdate), rec in sorted(records.items()):
        w = db.get(Worker, wid)
        ci = rec["check_in"]
        co = rec["check_out"]
        p = db.get(Project, ci.project_id) if ci else None
        rows.append({
            "worker_code": w.code if w else "?",
            "worker_name": w.name if w else "?",
            "project_code": p.code if p else "?",
            "project_name": p.name if p else "?",
            "date": wdate,
            "check_in": ci.server_time.isoformat() if ci else None,
            "check_out": co.server_time.isoformat() if co else None,
            "status": (ci.status.value if hasattr(ci.status, 'value') else str(ci.status)) if ci else None,
            "has_signature": bool(ci and ci.signature),
        })

    return _envelope(
        {
            "date_from": str(date_from),
            "date_to": str(date_to),
            "total_records": len(rows),
            "rows": rows,
        },
        request,
    )


@router.get("/reports/attendance/export")
def export_attendance(
    request: Request,
    date_from: date = Query(...),
    date_to: date = Query(...),
    project_id: str | None = Query(None),
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Export laporan kehadiran ke CSV/Excel."""
    # Reuse report logic
    tid = ctx.membership.tenant_id
    q = (
        select(AttendanceEvent)
        .where(
            AttendanceEvent.tenant_id == tid,
            AttendanceEvent.work_date >= date_from,
            AttendanceEvent.work_date <= date_to,
            AttendanceEvent.status.in_([AttendanceStatus.VALID, AttendanceStatus.REVIEW]),
        )
        .order_by(AttendanceEvent.work_date, AttendanceEvent.server_time)
    )
    if project_id:
        q = q.where(AttendanceEvent.project_id == project_id)
    events = db.scalars(q).all()

    from collections import defaultdict
    records = defaultdict(lambda: {"check_in": None, "check_out": None})
    for ev in events:
        key = (ev.worker_id, str(ev.work_date))
        etype = ev.event_type.value if hasattr(ev.event_type, 'value') else str(ev.event_type)
        if etype == "CHECK_IN":
            records[key]["check_in"] = ev
        elif etype == "CHECK_OUT":
            records[key]["check_out"] = ev

    rows = []
    for (wid, wdate), rec in sorted(records.items()):
        w = db.get(Worker, wid)
        ci = rec["check_in"]
        co = rec["check_out"]
        p = db.get(Project, ci.project_id) if ci else None
        ci_time = ci.server_time.astimezone(timezone.utc).strftime("%H:%M") if ci else ""
        co_time = co.server_time.astimezone(timezone.utc).strftime("%H:%M") if co else ""
        status_val = (ci.status.value if hasattr(ci.status, 'value') else str(ci.status)) if ci else ""
        rows.append([
            w.code if w else "", w.name if w else "",
            p.code if p else "", p.name if p else "",
            wdate, ci_time, co_time, status_val,
        ])

    if format == "csv":
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Kode", "Nama", "Proyek", "Nama Proyek", "Tanggal", "Jam Masuk", "Jam Keluar", "Status"])
        writer.writerows(rows)
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=laporan_absen_{date_from}_{date_to}.csv"},
        )
    else:
        # XLSX
        try:
            from openpyxl import Workbook
        except ImportError:
            raise HTTPException(500, detail={"code": "XLSX_NOT_INSTALLED"})
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Absen"
        ws.append(["Kode", "Nama", "Proyek", "Nama Proyek", "Tanggal", "Jam Masuk", "Jam Keluar", "Status"])
        for row in rows:
            ws.append(row)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=laporan_absen_{date_from}_{date_to}.xlsx"},
        )


# ──────────────────────────────────────────────
# KOREKSI ABSENSI + AUDIT TRAIL
# ──────────────────────────────────────────────

class AttendanceCorrection(BaseModel):
    new_check_in: str | None = Field(None, description="Jam masuk baru (HH:MM)")
    new_check_out: str | None = Field(None, description="Jam keluar baru (HH:MM)")
    reason: str = Field(min_length=5, max_length=500, description="Alasan koreksi")


@router.post("/corrections/{event_id}")
def correct_attendance(
    event_id: str,
    body: AttendanceCorrection,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Koreksi jam absen. Data asli tetap tersimpan di audit_events."""

    # Only OWNER and ADMIN can correct
    if ctx.membership.role not in (RoleCode.OWNER, RoleCode.ADMIN):
        raise HTTPException(403, detail={"code": "INSUFFICIENT_ROLE"})

    ev = db.scalar(
        select(AttendanceEvent).where(
            AttendanceEvent.id == event_id,
            AttendanceEvent.tenant_id == ctx.membership.tenant_id,
        )
    )
    if not ev:
        raise HTTPException(404, detail={"code": "EVENT_NOT_FOUND"})

    # Save original data to audit trail
    original_data = {
        "event_id": ev.id,
        "worker_id": ev.worker_id,
        "event_type": ev.event_type.value,
        "server_time": ev.server_time.isoformat(),
        "captured_at_client": ev.captured_at_client.isoformat(),
        "latitude": ev.latitude,
        "longitude": ev.longitude,
        "status": ev.status.value,
        "reason_code": ev.reason_code,
    }

    # Parse new times
    from datetime import time as dtime

    if body.new_check_in:
        try:
            h, m = map(int, body.new_check_in.split(":"))
            new_time = ev.server_time.replace(hour=h, minute=m, second=0, microsecond=0)
            ev.server_time = new_time
            ev.captured_at_client = new_time
        except (ValueError, AttributeError):
            raise HTTPException(400, detail={"code": "INVALID_TIME_FORMAT"})

    if body.new_check_out:
        try:
            h, m = map(int, body.new_check_out.split(":"))
            new_time = ev.server_time.replace(hour=h, minute=m, second=0, microsecond=0)
            ev.server_time = new_time
            ev.captured_at_client = new_time
        except (ValueError, AttributeError):
            raise HTTPException(400, detail={"code": "INVALID_TIME_FORMAT"})

    # If was REJECTED, change to VALID after correction
    if ev.status == AttendanceStatus.REJECTED:
        ev.status = AttendanceStatus.VALID

    # Write audit event
    audit = AuditEvent(
        id=str(__import__("uuid").uuid4()),
        tenant_id=ctx.membership.tenant_id,
        actor_user_id=ctx.user.id,
        action="ATTENDANCE_CORRECTION",
        entity_type="attendance_event",
        entity_id=ev.id,
        reason=body.reason,
        correlation_id=request.state.correlation_id,
        payload_json=json.dumps({
            "original": original_data,
            "corrected_to": {
                "server_time": ev.server_time.isoformat(),
                "new_check_in": body.new_check_in,
                "new_check_out": body.new_check_out,
            },
        }),
    )
    db.add(audit)
    db.commit()

    return _envelope({
        "event_id": ev.id,
        "corrected_by": ctx.user.display_name,
        "reason": body.reason,
        "original_time": original_data["server_time"],
        "new_time": ev.server_time.isoformat(),
        "audit_id": audit.id,
    }, request)


@router.get("/corrections/history")
def correction_history(
    request: Request,
    event_id: str | None = Query(None),
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Riwayat koreksi absensi."""

    q = (
        select(AuditEvent)
        .where(
            AuditEvent.tenant_id == ctx.membership.tenant_id,
            AuditEvent.action == "ATTENDANCE_CORRECTION",
        )
        .order_by(AuditEvent.created_at.desc())
    )
    if event_id:
        q = q.where(AuditEvent.entity_id == event_id)

    audits = db.scalars(q.limit(100)).all()
    result = []
    for a in audits:
        actor = db.get(User, a.actor_user_id) if a.actor_user_id else None
        result.append({
            "audit_id": a.id,
            "event_id": a.entity_id,
            "action": a.action,
            "reason": a.reason,
            "corrected_by": actor.display_name if actor else "?",
            "corrected_at": a.created_at.isoformat(),
            "payload": a.payload_json,
        })
    return _envelope(result, request)


# ──────────────────────────────────────────────
# LAPORAN LENGKAP HRD
# ──────────────────────────────────────────────

@router.get("/reports/monthly")
def monthly_report(
    request: Request,
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    project_id: str | None = Query(None),
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Laporan bulanan lengkap sesuai format HRD."""
    from calendar import monthrange
    from datetime import timedelta

    tid = ctx.membership.tenant_id
    _, last_day = monthrange(year, month)
    date_from = date(year, month, 1)
    date_to = date(year, month, last_day)

    # Get all workers
    workers = db.scalars(
        select(Worker).where(Worker.tenant_id == tid, Worker.is_active.is_(True)).order_by(Worker.code)
    ).all()

    # Get all attendance events for the month
    events = db.scalars(
        select(AttendanceEvent).where(
            AttendanceEvent.tenant_id == tid,
            AttendanceEvent.work_date >= date_from,
            AttendanceEvent.work_date <= date_to,
            AttendanceEvent.status.in_([AttendanceStatus.VALID, AttendanceStatus.REVIEW]),
        )
    ).all()

    # Group events by worker + date
    from collections import defaultdict
    event_map = defaultdict(lambda: {"check_in": None, "check_out": None})
    for ev in events:
        key = (ev.worker_id, str(ev.work_date))
        etype = ev.event_type.value if hasattr(ev.event_type, "value") else str(ev.event_type)
        if etype == "CHECK_IN":
            event_map[key]["check_in"] = ev
        elif etype == "CHECK_OUT":
            event_map[key]["check_out"] = ev

    # Get work schedules for the month
    schedules = db.scalars(
        select(WorkSchedule).where(
            WorkSchedule.tenant_id == tid,
            WorkSchedule.work_date >= date_from,
            WorkSchedule.work_date <= date_to,
        )
    ).all()
    schedule_map = {(s.worker_id, str(s.work_date)): s for s in schedules}

    # Get projects
    projects = {p.id: p for p in db.scalars(select(Project).where(Project.tenant_id == tid)).all()}

    # Build per-worker summary
    rows = []
    for w in workers:
        total_work_days = 0
        total_present = 0
        total_late = 0
        total_early_leave = 0
        total_absent = 0
        total_hours = 0.0
        daily_records = []

        current = date_from
        while current <= date_to:
            sched = schedule_map.get((w.id, str(current)))
            is_workday = sched.is_working_day if sched else (current.weekday() < 5)  # default Mon-Fri

            if is_workday:
                total_work_days += 1
                ev_data = event_map.get((w.id, str(current)))
                ci = ev_data["check_in"] if ev_data else None
                co = ev_data["check_out"] if ev_data else None

                if ci:
                    total_present += 1
                    # Calculate hours
                    if co:
                        delta = co.server_time - ci.server_time
                        hours = delta.total_seconds() / 3600
                        total_hours += hours
                    # Check late (after 08:00 + grace)
                    ci_time = ci.server_time.astimezone(timezone.utc)
                    if ci_time.hour > 8 or (ci_time.hour == 8 and ci_time.minute > 0):
                        total_late += 1
                    daily_records.append({
                        "date": str(current),
                        "check_in": ci.server_time.isoformat() if ci else None,
                        "check_out": co.server_time.isoformat() if co else None,
                        "status": "HADIR",
                        "hours": round(hours, 2) if co else None,
                    })
                else:
                    total_absent += 1
                    daily_records.append({
                        "date": str(current),
                        "check_in": None,
                        "check_out": None,
                        "status": "TANPA_KETERANGAN",
                        "hours": None,
                    })

            current += timedelta(days=1)

        # Get project assignment
        assignment = db.scalar(
            select(Assignment).where(Assignment.worker_id == w.id).limit(1)
        )
        proj = projects.get(assignment.project_id) if assignment else None

        rows.append({
            "worker_code": w.code,
            "worker_name": w.name,
            "project_code": proj.code if proj else "-",
            "project_name": proj.name if proj else "-",
            "total_work_days": total_work_days,
            "total_present": total_present,
            "total_late": total_late,
            "total_early_leave": total_early_leave,
            "total_absent": total_absent,
            "total_hours": round(total_hours, 2),
            "daily": daily_records,
        })

    return _envelope({
        "year": year,
        "month": month,
        "date_from": str(date_from),
        "date_to": str(date_to),
        "total_workers": len(rows),
        "rows": rows,
    }, request)


@router.get("/reports/monthly/export")
def export_monthly_report(
    request: Request,
    year: int = Query(...),
    month: int = Query(..., ge=1, le=12),
    project_id: str | None = Query(None),
    format: str = Query("xlsx", pattern="^(csv|xlsx)$"),
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Export laporan bulanan ke CSV/Excel."""
    # Reuse monthly report logic
    from calendar import monthrange
    from datetime import timedelta

    tid = ctx.membership.tenant_id
    _, last_day = monthrange(year, month)
    date_from = date(year, month, 1)
    date_to = date(year, month, last_day)

    workers = db.scalars(
        select(Worker).where(Worker.tenant_id == tid, Worker.is_active.is_(True)).order_by(Worker.code)
    ).all()

    events = db.scalars(
        select(AttendanceEvent).where(
            AttendanceEvent.tenant_id == tid,
            AttendanceEvent.work_date >= date_from,
            AttendanceEvent.work_date <= date_to,
            AttendanceEvent.status.in_([AttendanceStatus.VALID, AttendanceStatus.REVIEW]),
        )
    ).all()

    from collections import defaultdict
    event_map = defaultdict(lambda: {"check_in": None, "check_out": None})
    for ev in events:
        key = (ev.worker_id, str(ev.work_date))
        etype = ev.event_type.value if hasattr(ev.event_type, "value") else str(ev.event_type)
        if etype == "CHECK_IN":
            event_map[key]["check_in"] = ev
        elif etype == "CHECK_OUT":
            event_map[key]["check_out"] = ev

    schedules = db.scalars(
        select(WorkSchedule).where(
            WorkSchedule.tenant_id == tid,
            WorkSchedule.work_date >= date_from,
            WorkSchedule.work_date <= date_to,
        )
    ).all()
    schedule_map = {(s.worker_id, str(s.work_date)): s for s in schedules}

    projects = {p.id: p for p in db.scalars(select(Project).where(Project.tenant_id == tid)).all()}

    header = ["Kode", "Nama", "Proyek", "Total Hari Kerja", "Hadir", "Terlambat", "Alpha", "Total Jam Kerja"]
    rows = []
    for w in workers:
        total_work_days = 0
        total_present = 0
        total_late = 0
        total_absent = 0
        total_hours = 0.0

        current = date_from
        while current <= date_to:
            sched = schedule_map.get((w.id, str(current)))
            is_workday = sched.is_working_day if sched else (current.weekday() < 5)
            if is_workday:
                total_work_days += 1
                ev_data = event_map.get((w.id, str(current)))
                ci = ev_data["check_in"] if ev_data else None
                co = ev_data["check_out"] if ev_data else None
                if ci:
                    total_present += 1
                    if co:
                        delta = co.server_time - ci.server_time
                        total_hours += delta.total_seconds() / 3600
                    ci_time = ci.server_time.astimezone(timezone.utc)
                    if ci_time.hour > 8 or (ci_time.hour == 8 and ci_time.minute > 0):
                        total_late += 1
                else:
                    total_absent += 1
            current += timedelta(days=1)

        assignment = db.scalar(select(Assignment).where(Assignment.worker_id == w.id).limit(1))
        proj = projects.get(assignment.project_id) if assignment else None

        rows.append([
            w.code, w.name, proj.code if proj else "-",
            total_work_days, total_present, total_late, total_absent,
            round(total_hours, 2),
        ])

    if format == "csv":
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(header)
        writer.writerows(rows)
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=laporan_bulanan_{year}_{month:02d}.csv"},
        )
    else:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = f"Laporan {year}-{month:02d}"

        # Title
        ws.merge_cells("A1:H1")
        ws["A1"] = f"LAPORAN KEHADIRAN BULANAN — {year}/{month:02d}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Header
        header_fill = PatternFill(start_color="102A43", end_color="102A43", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        for col, h in enumerate(header, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data
        for r, row in enumerate(rows, 4):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)

        # Auto-width (skip merged cells)
        for col_idx in range(1, len(header) + 1):
            col_letter = ws.cell(row=3, column=col_idx).column_letter
            max_len = len(str(header[col_idx - 1]))
            for row_idx in range(4, len(rows) + 4):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 2

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=laporan_bulanan_{year}_{month:02d}.xlsx"},
        )


# ──────────────────────────────────────────────
# ROLE-BASED ACCESS CHECK
# ──────────────────────────────────────────────

def require_role(*roles: RoleCode):
    """Dependency factory: require specific roles."""
    def checker(ctx: RequestContext = Depends(tenant_context)) -> RequestContext:
        if ctx.membership.role not in roles:
            raise HTTPException(403, detail={"code": "INSUFFICIENT_ROLE", "required": [r.value for r in roles]})
        return ctx
    return checker


# Permission matrix
PERMISSIONS = {
    "manage_workers": (RoleCode.OWNER, RoleCode.ADMIN),
    "manage_schedules": (RoleCode.OWNER, RoleCode.ADMIN),
    "approve_review": (RoleCode.OWNER, RoleCode.SUPERVISOR),
    "correct_attendance": (RoleCode.OWNER, RoleCode.ADMIN),
    "view_reports": (RoleCode.OWNER, RoleCode.ADMIN, RoleCode.SUPERVISOR),
    "export_reports": (RoleCode.OWNER, RoleCode.ADMIN),
    "manage_holidays": (RoleCode.OWNER, RoleCode.ADMIN),
    "manage_projects": (RoleCode.OWNER, RoleCode.ADMIN),
}


@router.get("/permissions")
def get_permissions(
    request: Request,
    ctx: RequestContext = Depends(tenant_context),
):
    """Dapatkan permission user saat ini."""
    role = ctx.membership.role
    user_perms = {k: role in v for k, v in PERMISSIONS.items()}
    return _envelope({
        "role": role.value,
        "permissions": user_perms,
    }, request)


# ──────────────────────────────────────────────
# IZIN / SAKIT / CUTI / DINAS LUAR
# ──────────────────────────────────────────────

class LeaveCreate(BaseModel):
    worker_id: str
    leave_type: str = Field(pattern="^(IZIN|SAKIT|CUTI|DINAS_LUAR)$")
    date_from: date
    date_to: date
    reason: str = Field(min_length=3, max_length=500)


class LeaveDecision(BaseModel):
    decision: str = Field(pattern="^(APPROVED|REJECTED)$")
    note: str | None = None


@router.get("/leaves")
def list_leaves(
    request: Request,
    status: str | None = Query(None),
    worker_id: str | None = Query(None),
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """List pengajuan izin/sakit/cuti/dinas."""
    from app.models import LeaveRequest, LeaveStatus

    q = select(LeaveRequest).where(LeaveRequest.tenant_id == ctx.membership.tenant_id)
    if status:
        q = q.where(LeaveRequest.status == LeaveStatus(status))
    if worker_id:
        q = q.where(LeaveRequest.worker_id == worker_id)
    q = q.order_by(LeaveRequest.created_at.desc())

    leaves = db.scalars(q.limit(100)).all()
    result = []
    for lv in leaves:
        w = db.get(Worker, lv.worker_id)
        reviewer = db.get(User, lv.reviewed_by) if lv.reviewed_by else None
        result.append({
            "id": lv.id,
            "worker_code": w.code if w else "?",
            "worker_name": w.name if w else "?",
            "leave_type": lv.leave_type.value,
            "date_from": str(lv.date_from),
            "date_to": str(lv.date_to),
            "reason": lv.reason,
            "status": lv.status.value,
            "reviewed_by": reviewer.display_name if reviewer else None,
            "reviewed_at": lv.reviewed_at.isoformat() if lv.reviewed_at else None,
            "review_note": lv.review_note,
            "created_at": lv.created_at.isoformat(),
        })
    return _envelope(result, request)


@router.post("/leaves", status_code=201)
def create_leave(
    body: LeaveCreate,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Buat pengajuan izin/sakit/cuti/dinas (oleh admin untuk karyawan)."""
    from app.models import LeaveRequest, LeaveStatus, LeaveType

    worker = db.scalar(
        select(Worker).where(
            Worker.id == body.worker_id,
            Worker.tenant_id == ctx.membership.tenant_id,
        )
    )
    if not worker:
        raise HTTPException(404, detail={"code": "WORKER_NOT_FOUND"})

    leave = LeaveRequest(
        id=str(__import__("uuid").uuid4()),
        tenant_id=ctx.membership.tenant_id,
        worker_id=body.worker_id,
        leave_type=LeaveType(body.leave_type),
        date_from=body.date_from,
        date_to=body.date_to,
        reason=body.reason,
        status=LeaveStatus.APPROVED,  # Admin langsung approve
    )
    db.add(leave)
    db.commit()

    return _envelope({
        "id": leave.id,
        "worker_code": worker.code,
        "leave_type": leave.leave_type.value,
        "date_from": str(leave.date_from),
        "date_to": str(leave.date_to),
        "status": leave.status.value,
    }, request)


@router.post("/leaves/{leave_id}/decide")
def decide_leave(
    leave_id: str,
    body: LeaveDecision,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Approve atau reject pengajuan izin."""
    from app.models import LeaveRequest, LeaveStatus

    leave = db.scalar(
        select(LeaveRequest).where(
            LeaveRequest.id == leave_id,
            LeaveRequest.tenant_id == ctx.membership.tenant_id,
        )
    )
    if not leave:
        raise HTTPException(404, detail={"code": "LEAVE_NOT_FOUND"})
    if leave.status != LeaveStatus.PENDING:
        raise HTTPException(409, detail={"code": "NOT_PENDING"})

    leave.status = LeaveStatus(body.decision)
    leave.reviewed_by = ctx.user.id
    leave.reviewed_at = datetime.now(timezone.utc)
    leave.review_note = body.note
    db.commit()

    return _envelope({
        "id": leave.id,
        "status": leave.status.value,
        "reviewed_by": ctx.user.display_name,
    }, request)


# ──────────────────────────────────────────────
# LEMBUR (OVERTIME)
# ──────────────────────────────────────────────

class OvertimeCreate(BaseModel):
    worker_id: str
    work_date: date
    start_time: str = Field(description="HH:MM")
    end_time: str = Field(description="HH:MM")
    reason: str = Field(min_length=3, max_length=500)
    is_holiday: bool = False


class OvertimeDecision(BaseModel):
    decision: str = Field(pattern="^(APPROVED|REJECTED)$")
    note: str | None = None


@router.get("/overtimes")
def list_overtimes(
    request: Request,
    status: str | None = Query(None),
    worker_id: str | None = Query(None),
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """List pengajuan lembur."""
    from app.models import OvertimeRequest, OvertimeStatus

    q = select(OvertimeRequest).where(OvertimeRequest.tenant_id == ctx.membership.tenant_id)
    if status:
        q = q.where(OvertimeRequest.status == OvertimeStatus(status))
    if worker_id:
        q = q.where(OvertimeRequest.worker_id == worker_id)
    q = q.order_by(OvertimeRequest.created_at.desc())

    overtimes = db.scalars(q.limit(100)).all()
    result = []
    for ot in overtimes:
        w = db.get(Worker, ot.worker_id)
        reviewer = db.get(User, ot.reviewed_by) if ot.reviewed_by else None
        result.append({
            "id": ot.id,
            "worker_code": w.code if w else "?",
            "worker_name": w.name if w else "?",
            "work_date": str(ot.work_date),
            "start_time": ot.start_time.isoformat(),
            "end_time": ot.end_time.isoformat(),
            "hours": ot.hours,
            "reason": ot.reason,
            "is_holiday": ot.is_holiday,
            "status": ot.status.value,
            "reviewed_by": reviewer.display_name if reviewer else None,
            "reviewed_at": ot.reviewed_at.isoformat() if ot.reviewed_at else None,
            "review_note": ot.review_note,
            "created_at": ot.created_at.isoformat(),
        })
    return _envelope(result, request)


@router.post("/overtimes", status_code=201)
def create_overtime(
    body: OvertimeCreate,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Buat pengajuan lembur."""
    from app.models import OvertimeRequest, OvertimeStatus

    worker = db.scalar(
        select(Worker).where(
            Worker.id == body.worker_id,
            Worker.tenant_id == ctx.membership.tenant_id,
        )
    )
    if not worker:
        raise HTTPException(404, detail={"code": "WORKER_NOT_FOUND"})

    # Parse times
    sh, sm = map(int, body.start_time.split(":"))
    eh, em = map(int, body.end_time.split(":"))
    start_dt = datetime(body.work_date.year, body.work_date.month, body.work_date.day, sh, sm, tzinfo=timezone.utc)
    end_dt = datetime(body.work_date.year, body.work_date.month, body.work_date.day, eh, em, tzinfo=timezone.utc)
    hours = (end_dt - start_dt).total_seconds() / 3600

    ot = OvertimeRequest(
        id=str(__import__("uuid").uuid4()),
        tenant_id=ctx.membership.tenant_id,
        worker_id=body.worker_id,
        work_date=body.work_date,
        start_time=start_dt,
        end_time=end_dt,
        hours=round(hours, 2),
        reason=body.reason,
        is_holiday=body.is_holiday,
        status=OvertimeStatus.PENDING,
    )
    db.add(ot)
    db.commit()

    return _envelope({
        "id": ot.id,
        "worker_code": worker.code,
        "hours": ot.hours,
        "status": ot.status.value,
    }, request)


@router.post("/overtimes/{ot_id}/decide")
def decide_overtime(
    ot_id: str,
    body: OvertimeDecision,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Approve atau reject pengajuan lembur."""
    from app.models import OvertimeRequest, OvertimeStatus

    ot = db.scalar(
        select(OvertimeRequest).where(
            OvertimeRequest.id == ot_id,
            OvertimeRequest.tenant_id == ctx.membership.tenant_id,
        )
    )
    if not ot:
        raise HTTPException(404, detail={"code": "OVERTIME_NOT_FOUND"})
    if ot.status != OvertimeStatus.PENDING:
        raise HTTPException(409, detail={"code": "NOT_PENDING"})

    ot.status = OvertimeStatus(body.decision)
    ot.reviewed_by = ctx.user.id
    ot.reviewed_at = datetime.now(timezone.utc)
    ot.review_note = body.note
    db.commit()

    return _envelope({
        "id": ot.id,
        "status": ot.status.value,
        "reviewed_by": ctx.user.display_name,
    }, request)
