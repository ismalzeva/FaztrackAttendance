"""Schedule Management — Opsi B: Template + Override + Holiday Calendar.

Models:
- ScheduleTemplate: pola kerja (Sen-Sab, Sen-Jum, Sales Flexible, dll)
- EmployeeSchedule: assign template ke karyawan
- ScheduleOverride: override jadwal per karyawan per tanggal
- HolidayCalendar: tanggal merah nasional + libur perusahaan

Logic:
- Jadwal efektif = template + override + holiday check
- HRD bisa bulk assign template via Excel
- Override per tanggal untuk kasus khusus
"""
import io
from datetime import date, datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import select, delete, and_
from sqlalchemy.orm import Session

from app.database import get_db
from app.dependencies import RequestContext, admin_context
from app.models import Worker, Tenant

# We'll define the models here and create tables on first use
from sqlalchemy import Column, String, Boolean, Date, DateTime, Integer, ForeignKey, UniqueConstraint, Text
from sqlalchemy.orm import relationship
from app.database import Base
from app.models import uid, now


class ScheduleTemplate(Base):
    __tablename__ = "schedule_templates"
    id = Column(String(36), primary_key=True, default=uid)
    tenant_id = Column(String(36), ForeignKey("tenants.id"), index=True, nullable=False)
    name = Column(String(100), nullable=False)
    description = Column(Text, nullable=True)
    # Work days: 0=Mon, 1=Tue, 2=Wed, 3=Thu, 4=Fri, 5=Sat, 6=Sun
    work_days = Column(String(20), nullable=False, default="0,1,2,3,4,5")  # comma-separated day indices
    work_start = Column(String(5), nullable=False, default="08:00")  # HH:MM
    work_end = Column(String(5), nullable=False, default="17:00")  # HH:MM
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime(timezone=True), default=now)


class EmployeeSchedule(Base):
    __tablename__ = "employee_schedules"
    __table_args__ = (UniqueConstraint("tenant_id", "worker_id", name="uq_emp_schedule_worker"),)
    id = Column(String(36), primary_key=True, default=uid)
    tenant_id = Column(String(36), ForeignKey("tenants.id"), index=True, nullable=False)
    worker_id = Column(String(36), ForeignKey("workers.id"), index=True, nullable=False)
    template_id = Column(String(36), ForeignKey("schedule_templates.id"), nullable=False)
    effective_from = Column(Date, nullable=True)  # null = immediate
    effective_until = Column(Date, nullable=True)  # null = indefinite
    created_at = Column(DateTime(timezone=True), default=now)


class ScheduleOverride(Base):
    __tablename__ = "schedule_overrides"
    __table_args__ = (UniqueConstraint("tenant_id", "worker_id", "override_date", name="uq_schedule_override"),)
    id = Column(String(36), primary_key=True, default=uid)
    tenant_id = Column(String(36), ForeignKey("tenants.id"), index=True, nullable=False)
    worker_id = Column(String(36), ForeignKey("workers.id"), index=True, nullable=False)
    override_date = Column(Date, nullable=False, index=True)
    is_workday = Column(Boolean, nullable=False)  # True = masuk, False = libur
    reason = Column(String(200), nullable=True)  # cth: "Lembur", "Ganti libur", "Dinas luar"
    created_at = Column(DateTime(timezone=True), default=now)


class HolidayCalendar(Base):
    __tablename__ = "holiday_calendar"
    __table_args__ = (UniqueConstraint("tenant_id", "holiday_date", name="uq_holiday_date"),)
    id = Column(String(36), primary_key=True, default=uid)
    tenant_id = Column(String(36), ForeignKey("tenants.id"), index=True, nullable=False)
    holiday_date = Column(Date, nullable=False, index=True)
    name = Column(String(200), nullable=False)
    holiday_type = Column(String(20), nullable=False, default="national")  # national / company
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime(timezone=True), default=now)


router = APIRouter(prefix="/api/v1/schedule", tags=["schedule-management"])


def _envelope(data, request: Request):
    return {
        "data": data,
        "meta": {
            "correlation_id": request.state.correlation_id,
            "server_time": datetime.now(timezone.utc).isoformat(),
            "version": "v1",
        },
    }


# ──────────────────────────────────────────────
# TEMPLATE CRUD
# ──────────────────────────────────────────────

class TemplateCreate(BaseModel):
    name: str = Field(min_length=1, max_length=100)
    description: str | None = None
    work_days: str = Field(default="0,1,2,3,4,5")  # comma-separated
    work_start: str = Field(default="08:00")
    work_end: str = Field(default="17:00")


class TemplateUpdate(BaseModel):
    name: str | None = None
    description: str | None = None
    work_days: str | None = None
    work_start: str | None = None
    work_end: str | None = None
    is_active: bool | None = None


@router.get("/templates")
def list_templates(
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """List semua template jadwal."""
    tid = ctx.membership.tenant_id
    templates = db.scalars(
        select(ScheduleTemplate).where(ScheduleTemplate.tenant_id == tid).order_by(ScheduleTemplate.name)
    ).all()
    return _envelope(
        [
            {
                "id": t.id,
                "name": t.name,
                "description": t.description,
                "work_days": t.work_days,
                "work_start": t.work_start,
                "work_end": t.work_end,
                "is_active": t.is_active,
            }
            for t in templates
        ],
        request,
    )


@router.post("/templates", status_code=201)
def create_template(
    body: TemplateCreate,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Buat template jadwal baru."""
    tid = ctx.membership.tenant_id
    t = ScheduleTemplate(
        tenant_id=tid,
        name=body.name,
        description=body.description,
        work_days=body.work_days,
        work_start=body.work_start,
        work_end=body.work_end,
    )
    db.add(t)
    db.commit()
    return _envelope({"id": t.id, "name": t.name}, request)


@router.patch("/templates/{template_id}")
def update_template(
    template_id: str,
    body: TemplateUpdate,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Edit template jadwal."""
    t = db.scalar(
        select(ScheduleTemplate).where(
            ScheduleTemplate.id == template_id,
            ScheduleTemplate.tenant_id == ctx.membership.tenant_id,
        )
    )
    if not t:
        raise HTTPException(404, detail={"code": "TEMPLATE_NOT_FOUND"})
    for field in ["name", "description", "work_days", "work_start", "work_end", "is_active"]:
        val = getattr(body, field, None)
        if val is not None:
            setattr(t, field, val)
    db.commit()
    return _envelope({"id": t.id, "name": t.name}, request)


@router.delete("/templates/{template_id}")
def delete_template(
    template_id: str,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Hapus template (jika tidak dipakai karyawan)."""
    t = db.scalar(
        select(ScheduleTemplate).where(
            ScheduleTemplate.id == template_id,
            ScheduleTemplate.tenant_id == ctx.membership.tenant_id,
        )
    )
    if not t:
        raise HTTPException(404, detail={"code": "TEMPLATE_NOT_FOUND"})
    # Check if any employee uses this template
    count = db.scalar(
        select(EmployeeSchedule).where(EmployeeSchedule.template_id == template_id)
    )
    if count:
        raise HTTPException(409, detail={"code": "TEMPLATE_IN_USE", "message": "Template masih dipakai karyawan"})
    db.delete(t)
    db.commit()
    return _envelope({"deleted": True}, request)


# ──────────────────────────────────────────────
# ASSIGN TEMPLATE KE KARYAWAN
# ──────────────────────────────────────────────

class AssignTemplate(BaseModel):
    worker_id: str
    template_id: str
    effective_from: date | None = None
    effective_until: date | None = None


class BulkAssign(BaseModel):
    assignments: list[AssignTemplate]


@router.get("/assignments")
def list_assignments(
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """List semua assignment template ke karyawan."""
    tid = ctx.membership.tenant_id
    rows = db.scalars(
        select(EmployeeSchedule).where(EmployeeSchedule.tenant_id == tid)
    ).all()
    result = []
    for r in rows:
        w = db.get(Worker, r.worker_id)
        t = db.get(ScheduleTemplate, r.template_id)
        result.append({
            "id": r.id,
            "worker_id": r.worker_id,
            "worker_code": w.code if w else "?",
            "worker_name": w.name if w else "?",
            "template_id": r.template_id,
            "template_name": t.name if t else "?",
            "effective_from": str(r.effective_from) if r.effective_from else None,
            "effective_until": str(r.effective_until) if r.effective_until else None,
        })
    return _envelope(result, request)


@router.post("/assignments", status_code=201)
def assign_template(
    body: AssignTemplate,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Assign template ke satu karyawan."""
    tid = ctx.membership.tenant_id
    # Check exists
    existing = db.scalar(
        select(EmployeeSchedule).where(
            EmployeeSchedule.tenant_id == tid,
            EmployeeSchedule.worker_id == body.worker_id,
        )
    )
    if existing:
        existing.template_id = body.template_id
        existing.effective_from = body.effective_from
        existing.effective_until = body.effective_until
        db.commit()
        return _envelope({"id": existing.id, "updated": True}, request)
    es = EmployeeSchedule(
        tenant_id=tid,
        worker_id=body.worker_id,
        template_id=body.template_id,
        effective_from=body.effective_from,
        effective_until=body.effective_until,
    )
    db.add(es)
    db.commit()
    return _envelope({"id": es.id}, request)


@router.post("/assignments/bulk", status_code=201)
def bulk_assign(
    body: BulkAssign,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Bulk assign template ke banyak karyawan."""
    tid = ctx.membership.tenant_id
    created = 0
    updated = 0
    for a in body.assignments:
        existing = db.scalar(
            select(EmployeeSchedule).where(
                EmployeeSchedule.tenant_id == tid,
                EmployeeSchedule.worker_id == a.worker_id,
            )
        )
        if existing:
            existing.template_id = a.template_id
            existing.effective_from = a.effective_from
            existing.effective_until = a.effective_until
            updated += 1
        else:
            db.add(EmployeeSchedule(
                tenant_id=tid,
                worker_id=a.worker_id,
                template_id=a.template_id,
                effective_from=a.effective_from,
                effective_until=a.effective_until,
            ))
            created += 1
    db.commit()
    return _envelope({"created": created, "updated": updated}, request)


# ──────────────────────────────────────────────
# OVERRIDE JADWAL PER KARYAWAN
# ──────────────────────────────────────────────

class OverrideCreate(BaseModel):
    worker_id: str
    override_date: date
    is_workday: bool
    reason: str | None = None


@router.get("/overrides")
def list_overrides(
    request: Request,
    date_from: date = Query(...),
    date_to: date = Query(...),
    worker_id: str | None = Query(None),
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """List override jadwal dalam periode."""
    tid = ctx.membership.tenant_id
    q = select(ScheduleOverride).where(
        ScheduleOverride.tenant_id == tid,
        ScheduleOverride.override_date >= date_from,
        ScheduleOverride.override_date <= date_to,
    )
    if worker_id:
        q = q.where(ScheduleOverride.worker_id == worker_id)
    rows = db.scalars(q.order_by(ScheduleOverride.override_date)).all()
    result = []
    for r in rows:
        w = db.get(Worker, r.worker_id)
        result.append({
            "id": r.id,
            "worker_id": r.worker_id,
            "worker_code": w.code if w else "?",
            "worker_name": w.name if w else "?",
            "override_date": str(r.override_date),
            "is_workday": r.is_workday,
            "reason": r.reason,
        })
    return _envelope(result, request)


@router.post("/overrides", status_code=201)
def create_override(
    body: OverrideCreate,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Buat override jadwal (misal: karyawan masuk di hari libur)."""
    tid = ctx.membership.tenant_id
    existing = db.scalar(
        select(ScheduleOverride).where(
            ScheduleOverride.tenant_id == tid,
            ScheduleOverride.worker_id == body.worker_id,
            ScheduleOverride.override_date == body.override_date,
        )
    )
    if existing:
        existing.is_workday = body.is_workday
        existing.reason = body.reason
        db.commit()
        return _envelope({"id": existing.id, "updated": True}, request)
    o = ScheduleOverride(
        tenant_id=tid,
        worker_id=body.worker_id,
        override_date=body.override_date,
        is_workday=body.is_workday,
        reason=body.reason,
    )
    db.add(o)
    db.commit()
    return _envelope({"id": o.id}, request)


@router.delete("/overrides/{override_id}")
def delete_override(
    override_id: str,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Hapus override."""
    o = db.scalar(
        select(ScheduleOverride).where(
            ScheduleOverride.id == override_id,
            ScheduleOverride.tenant_id == ctx.membership.tenant_id,
        )
    )
    if not o:
        raise HTTPException(404, detail={"code": "OVERRIDE_NOT_FOUND"})
    db.delete(o)
    db.commit()
    return _envelope({"deleted": True}, request)


# ──────────────────────────────────────────────
# HOLIDAY CALENDAR
# ──────────────────────────────────────────────

class HolidayCreate(BaseModel):
    holiday_date: date
    name: str = Field(min_length=1, max_length=200)
    holiday_type: str = Field(default="national", pattern="^(national|company)$")


@router.get("/holidays")
def list_holidays(
    request: Request,
    year: int = Query(...),
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """List hari libur untuk tahun tertentu."""
    tid = ctx.membership.tenant_id
    rows = db.scalars(
        select(HolidayCalendar).where(
            HolidayCalendar.tenant_id == tid,
            HolidayCalendar.is_active.is_(True),
        ).order_by(HolidayCalendar.holiday_date)
    ).all()
    # Filter by year
    result = [
        {
            "id": r.id,
            "date": str(r.holiday_date),
            "name": r.name,
            "type": r.holiday_type,
        }
        for r in rows
        if r.holiday_date.year == year
    ]
    return _envelope(result, request)


@router.post("/holidays", status_code=201)
def create_holiday(
    body: HolidayCreate,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Tambah hari libur."""
    tid = ctx.membership.tenant_id
    existing = db.scalar(
        select(HolidayCalendar).where(
            HolidayCalendar.tenant_id == tid,
            HolidayCalendar.holiday_date == body.holiday_date,
        )
    )
    if existing:
        existing.name = body.name
        existing.holiday_type = body.holiday_type
        existing.is_active = True
        db.commit()
        return _envelope({"id": existing.id, "updated": True}, request)
    h = HolidayCalendar(
        tenant_id=tid,
        holiday_date=body.holiday_date,
        name=body.name,
        holiday_type=body.holiday_type,
    )
    db.add(h)
    db.commit()
    return _envelope({"id": h.id}, request)


@router.delete("/holidays/{holiday_id}")
def delete_holiday(
    holiday_id: str,
    request: Request,
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Hapus hari libur."""
    h = db.scalar(
        select(HolidayCalendar).where(
            HolidayCalendar.id == holiday_id,
            HolidayCalendar.tenant_id == ctx.membership.tenant_id,
        )
    )
    if not h:
        raise HTTPException(404, detail={"code": "HOLIDAY_NOT_FOUND"})
    db.delete(h)
    db.commit()
    return _envelope({"deleted": True}, request)


# ──────────────────────────────────────────────
# JADWAL EFEKTIF (query)
# ──────────────────────────────────────────────

@router.get("/effective")
def effective_schedule(
    request: Request,
    worker_id: str = Query(...),
    check_date: date = Query(...),
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Cek jadwal efektif karyawan untuk tanggal tertentu.
    
    Logic:
    1. Cek override → jika ada, pakai override
    2. Cek holiday → jika libur nasional/perusahaan, libur
    3. Cek template → hari dalam template
    """
    tid = ctx.membership.tenant_id
    
    # 1. Check override
    override = db.scalar(
        select(ScheduleOverride).where(
            ScheduleOverride.tenant_id == tid,
            ScheduleOverride.worker_id == worker_id,
            ScheduleOverride.override_date == check_date,
        )
    )
    if override:
        return _envelope({
            "worker_id": worker_id,
            "date": str(check_date),
            "is_workday": override.is_workday,
            "source": "override",
            "reason": override.reason,
        }, request)
    
    # 2. Check holiday
    holiday = db.scalar(
        select(HolidayCalendar).where(
            HolidayCalendar.tenant_id == tid,
            HolidayCalendar.holiday_date == check_date,
            HolidayCalendar.is_active.is_(True),
        )
    )
    if holiday:
        return _envelope({
            "worker_id": worker_id,
            "date": str(check_date),
            "is_workday": False,
            "source": "holiday",
            "holiday_name": holiday.name,
            "holiday_type": holiday.holiday_type,
        }, request)
    
    # 3. Check template
    emp_schedule = db.scalar(
        select(EmployeeSchedule).where(
            EmployeeSchedule.tenant_id == tid,
            EmployeeSchedule.worker_id == worker_id,
        )
    )
    if emp_schedule:
        template = db.get(ScheduleTemplate, emp_schedule.template_id)
        if template:
            day_of_week = check_date.weekday()  # 0=Mon, 6=Sun
            work_days = [int(d) for d in template.work_days.split(",") if d.strip()]
            is_workday = day_of_week in work_days
            return _envelope({
                "worker_id": worker_id,
                "date": str(check_date),
                "is_workday": is_workday,
                "source": "template",
                "template_name": template.name,
                "work_start": template.work_start,
                "work_end": template.work_end,
            }, request)
    
    # 4. No schedule found — default to workday (Mon-Sat)
    day_of_week = check_date.weekday()
    return _envelope({
        "worker_id": worker_id,
        "date": str(check_date),
        "is_workday": day_of_week < 6,  # Mon-Sat
        "source": "default",
    }, request)


# ──────────────────────────────────────────────
# IMPORT EXCEL — BULK ASSIGN TEMPLATE
# ──────────────────────────────────────────────

@router.post("/import/assign-templates")
async def import_assign_templates(
    request: Request,
    file: UploadFile = File(...),
    ctx: RequestContext = Depends(admin_context),
    db: Session = Depends(get_db),
):
    """Import Excel untuk bulk assign template ke karyawan.
    
    Format Excel:
    | Kode Karyawan | Template Name |
    |---------------|---------------|
    | EMP-001       | Senin-Sabtu   |
    | EMP-020       | Sales Flexible|
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, detail={"code": "XLSX_NOT_INSTALLED"})
    
    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    
    tid = ctx.membership.tenant_id
    
    # Build template map
    templates = db.scalars(
        select(ScheduleTemplate).where(ScheduleTemplate.tenant_id == tid)
    ).all()
    template_map = {t.name.lower(): t.id for t in templates}
    
    # Build worker map
    workers = db.scalars(
        select(Worker).where(Worker.tenant_id == tid, Worker.is_active.is_(True))
    ).all()
    worker_map = {w.code.lower(): w.id for w in workers}
    
    created = 0
    updated = 0
    errors = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        worker_code = str(row[0]).strip().lower()
        template_name = str(row[1]).strip().lower() if len(row) > 1 and row[1] else None
        
        if worker_code not in worker_map:
            errors.append(f"Row {row_idx}: Karyawan '{worker_code}' tidak ditemukan")
            continue
        if not template_name or template_name not in template_map:
            errors.append(f"Row {row_idx}: Template '{template_name}' tidak ditemukan")
            continue
        
        worker_id = worker_map[worker_code]
        template_id = template_map[template_name]
        
        existing = db.scalar(
            select(EmployeeSchedule).where(
                EmployeeSchedule.tenant_id == tid,
                EmployeeSchedule.worker_id == worker_id,
            )
        )
        if existing:
            existing.template_id = template_id
            updated += 1
        else:
            db.add(EmployeeSchedule(
                tenant_id=tid,
                worker_id=worker_id,
                template_id=template_id,
            ))
            created += 1
    
    db.commit()
    return _envelope({
        "created": created,
        "updated": updated,
        "errors": errors,
    }, request)
