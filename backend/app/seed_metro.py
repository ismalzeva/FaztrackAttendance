"""
seed_metro.py — Import Metro Mining simulation workbook into the database.
Usage: cd backend && python -m app.seed_metro
"""
import sys
import os
from datetime import date, datetime, time

import openpyxl
from sqlalchemy import create_engine
from sqlalchemy.orm import Session

from app.database import Base
from app.config import get_settings
from app.security import hash_password
from app.models import (
    Tenant, User, Membership, RoleCode,
    Worker, Site, Equipment, Role, Crew, Competency,
    ShiftTemplate, CheckpointPolicy, RosterPolicy, RuleVersion,
    RosterAssignment, EmployeeMeta,
    WorkStatus, SiteStatusEnum, ValidationStatus,
    SiteType, SiteStatus, EquipmentStatus, CompetencyStatus,
)
from app.rule_versioning import snapshot_rules

WORKBOOK = "/home/ubuntu/.hermes/cache/documents/doc_82b2f5cfb574_METRO_MINING_MASTER_DATA_SIMULATION_v1.0.xlsx"

TENANT_ID = "metro-mining-001"
TENANT_CODE = "metro-mining"
ADMIN_USER_ID = "metro-admin-001"


def _parse_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    return datetime.fromisoformat(s.replace(" ", "T").split("T")[0]).date()


def _parse_time(val) -> time | None:
    if val is None:
        return None
    if isinstance(val, time):
        return val
    if isinstance(val, datetime):
        return val.time()
    s = str(val).strip()
    if not s:
        return None
    # Handle HH:MM format
    parts = s.split(":")
    return time(int(parts[0]), int(parts[1]) if len(parts) > 1 else 0)


def _parse_bool(val) -> bool:
    if val is None:
        return False
    if isinstance(val, bool):
        return val
    s = str(val).strip().upper()
    return s in ("TRUE", "1", "YES", "Y")


def _clean_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _clean_int(val) -> int | None:
    if val is None:
        return None
    if isinstance(val, int):
        return val
    s = str(val).strip()
    if not s or s.upper() == "TBC":
        return None
    return int(float(s))


def seed(db: Session) -> dict:
    """Run full Metro Mining seed. Returns counts dict."""
    wb_path = WORKBOOK
    print(f"Loading workbook: {wb_path}")
    wb = openpyxl.load_workbook(wb_path, data_only=True)

    counts = {}

    # 1. Tenant
    tenant = db.query(Tenant).filter(Tenant.id == TENANT_ID).first()
    if not tenant:
        tenant = Tenant(id=TENANT_ID, code=TENANT_CODE, name="Metro Mining", timezone="Asia/Makassar")
        db.add(tenant)
        db.flush()
    counts["tenant"] = 1

    # 1b. Ensure tenant timezone
    if tenant.timezone != "Asia/Makassar":
        tenant.timezone = "Asia/Makassar"
        db.flush()

    # 2. Admin user
    admin = db.query(User).filter(User.id == ADMIN_USER_ID).first()
    if not admin:
        admin = User(id=ADMIN_USER_ID, login_id="admin@metro-mining.id", display_name="Metro Admin", password_hash=hash_password("metro2026"))
        db.add(admin)
        db.flush()
    if not db.query(Membership).filter(Membership.user_id == ADMIN_USER_ID, Membership.tenant_id == TENANT_ID).first():
        db.add(Membership(id="metro-mem-001", user_id=ADMIN_USER_ID, tenant_id=TENANT_ID, role=RoleCode.OWNER))
    counts["admin"] = 1

    # 3. Config → RosterPolicy
    ws = wb["Config"]
    policy_count = 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        key = _clean_str(row[0])
        if not key:
            continue
        val = _clean_str(row[1])
        dtype = _clean_str(row[2]) or "string"
        status = _clean_str(row[3]) or "TBC"
        notes = _clean_str(row[4])
        if not db.query(RosterPolicy).filter(RosterPolicy.tenant_id == TENANT_ID, RosterPolicy.policy_key == key).first():
            db.add(RosterPolicy(
                id=f"rp-{key}",
                tenant_id=TENANT_ID,
                policy_key=key,
                policy_value=val or "TBC",
                data_type=dtype,
                confirmation_status=status,
                notes=notes,
            ))
            policy_count += 1
    counts["policies"] = policy_count

    # 4. Sites
    ws = wb["Sites"]
    site_count = 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        sid = _clean_str(row[0])
        if not sid:
            continue
        site_type_str = _clean_str(row[3]) or "MINE_SITE"
        lat = row[4]
        lon = row[5]
        radius = row[6]
        site = db.query(Site).filter(Site.id == sid).first()
        if not site:
            db.add(Site(
                id=sid,
                tenant_id=TENANT_ID,
                site_code=sid,
                site_name=_clean_str(row[2]) or sid,
                site_type=SiteType(site_type_str),
                latitude=float(lat) if lat and str(lat).upper() != "TBC" else None,
                longitude=float(lon) if lon and str(lon).upper() != "TBC" else None,
                radius_m=int(float(radius)) if radius and str(radius).upper() not in ("TBC", "DUMMY", "SIMULATION") else None,
                status=SiteStatus.ACTIVE,
                effective_from=date(2026, 9, 1),
                notes=f"[SIMULATION/NON_PRODUCTION] {_clean_str(row[8]) or 'Geofence data TBC'}",
            ))
            site_count += 1
    counts["sites"] = site_count

    # 4b. Set site timezone + geofence TBC
    for site in db.query(Site).filter(Site.tenant_id == TENANT_ID).all():
        site.timezone = "Asia/Makassar"
        site.radius_m = None  # TBC until geofence data provided
        if site.notes and "[SIMULATION" not in (site.notes or ""):
            site.notes = f"[SIMULATION/NON_PRODUCTION] {site.notes or 'Geofence data TBC'}"
    db.flush()

    # 5. Shifts
    ws = wb["Shifts & Checkpoints"]
    shift_map = {}
    shift_count = 0
    for row in ws.iter_rows(min_row=5, max_row=6, values_only=True):
        sid = _clean_str(row[0])
        if not sid:
            continue
        shift = db.query(ShiftTemplate).filter(ShiftTemplate.id == sid).first()
        if not shift:
            shift = ShiftTemplate(
                id=sid,
                tenant_id=TENANT_ID,
                shift_code=sid,
                shift_name=_clean_str(row[1]) or sid,
                start_time=_parse_time(row[2]) or time(7, 0),
                end_time=_parse_time(row[3]) or time(19, 0),
                break_start=_parse_time(row[4]) or time(12, 0),
                break_end=_parse_time(row[5]) or time(13, 0),
                handover_start=_parse_time(row[6]) or time(18, 45),
                handover_end=_parse_time(row[7]) or time(19, 0),
                crosses_midnight=_parse_bool(row[8]),
            )
            db.add(shift)
            shift_count += 1
        shift_map[sid] = sid
    counts["shifts"] = shift_count

    # 6. Checkpoint policies
    cp_count = 0
    for row in ws.iter_rows(min_row=10, values_only=True):
        cp_type = _clean_str(row[0])
        if not cp_type:
            continue
        for shift_id in ["DAY", "NIGHT"]:
            col_offset = 0 if shift_id == "DAY" else 1
            cp = db.query(CheckpointPolicy).filter(
                CheckpointPolicy.tenant_id == TENANT_ID,
                CheckpointPolicy.checkpoint_type == cp_type,
                CheckpointPolicy.shift_id == shift_id,
            ).first()
            if not cp:
                db.add(CheckpointPolicy(
                    id=f"cp-{cp_type.lower()}-{shift_id.lower()}",
                    tenant_id=TENANT_ID,
                    checkpoint_type=cp_type,
                    shift_id=shift_id,
                    window_start_offset_min=0,
                    window_end_offset_min=0,
                    required_evidence=_clean_str(row[3]),
                    severity=_clean_str(row[4]) or "WARNING",
                ))
                cp_count += 1
    counts["checkpoint_policies"] = cp_count

    # 7. Roles
    ws = wb["Roles & Crews"]
    role_count = 0
    for row in ws.iter_rows(min_row=5, max_row=7, values_only=True):
        rid = _clean_str(row[0])
        if not rid:
            continue
        if not db.query(Role).filter(Role.id == rid).first():
            db.add(Role(
                id=rid,
                tenant_id=TENANT_ID,
                role_code=rid,
                role_name=_clean_str(row[1]) or rid,
                equipment_type_required=_clean_str(row[2]),
                status=_clean_str(row[3]) or "ACTIVE",
            ))
            role_count += 1
    counts["roles"] = role_count

    # 8. Crews
    crew_count = 0
    for row in ws.iter_rows(min_row=11, values_only=True):
        cid = _clean_str(row[0])
        if not cid:
            continue
        if not db.query(Crew).filter(Crew.id == cid).first():
            db.add(Crew(
                id=cid,
                tenant_id=TENANT_ID,
                crew_code=cid,
                crew_name=_clean_str(row[1]) or cid,
                onsite_cycle_anchor=date(2026, 9, 1),
                cycle_offset_days=_clean_int(row[2]) or 0,
            ))
            crew_count += 1
    counts["crews"] = crew_count

    # 9. Employees → Worker + EmployeeMeta
    ws = wb["Employees"]
    emp_count = 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        eid = _clean_str(row[0])
        if not eid:
            continue
        if not db.query(Worker).filter(Worker.id == eid).first():
            is_active = (_clean_str(row[5]) or "ACTIVE") == "ACTIVE"
            db.add(Worker(
                id=eid,
                tenant_id=TENANT_ID,
                code=_clean_str(row[1]) or eid,
                name=_clean_str(row[2]) or eid,
                is_active=is_active,
            ))
            # EmployeeMeta
            role_id = _clean_str(row[3])
            crew_id = _clean_str(row[4])
            eff_from = _parse_date(row[6]) or date(2026, 9, 1)
            eff_to = _parse_date(row[7])
            cycle_offset = _clean_int(row[8]) or 0
            db.add(EmployeeMeta(
                id=f"em-{eid}",
                tenant_id=TENANT_ID,
                worker_id=eid,
                employee_no=_clean_str(row[1]),
                role_id=role_id,
                crew_id=crew_id,
                effective_from=eff_from,
                effective_to=eff_to,
                cycle_offset_days=cycle_offset,
            ))
            emp_count += 1
    counts["employees"] = emp_count

    # 10. Equipment
    ws = wb["Equipment"]
    eq_count = 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        eqid = _clean_str(row[0])
        if not eqid:
            continue
        if not db.query(Equipment).filter(Equipment.id == eqid).first():
            db.add(Equipment(
                id=eqid,
                tenant_id=TENANT_ID,
                equipment_code=_clean_str(row[1]) or eqid,
                equipment_type=_clean_str(row[2]) or "UNKNOWN",
                status=EquipmentStatus(_clean_str(row[3]) or "ACTIVE"),
                effective_from=_parse_date(row[4]) or date(2026, 9, 1),
                effective_to=_parse_date(row[5]),
                notes=_clean_str(row[6]),
            ))
            eq_count += 1
    counts["equipment"] = eq_count

    # 11. Competencies
    ws = wb["Competencies"]
    comp_count = 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        cid = _clean_str(row[0])
        if not cid:
            continue
        if not db.query(Competency).filter(Competency.id == cid).first():
            db.add(Competency(
                id=cid,
                tenant_id=TENANT_ID,
                competency_code=cid,
                employee_id=_clean_str(row[1]) or cid,
                equipment_type=_clean_str(row[2]) or "UNKNOWN",
                certification_no=_clean_str(row[3]),
                valid_from=_parse_date(row[4]) or date(2026, 1, 1),
                valid_to=_parse_date(row[5]),
                status=CompetencyStatus(_clean_str(row[6]) or "VALID"),
                source=_clean_str(row[7]),
                notes=_clean_str(row[8]),
            ))
            comp_count += 1
    counts["competencies"] = comp_count

    # 13. Snapshot rules (BEFORE roster so we have rule_version_id)
    rv = snapshot_rules(db, TENANT_ID, "METRO-RULE-v0.1", date(2026, 9, 1))
    counts["rule_version"] = 1

    # 12. Roster Simulation
    ws = wb["Roster Simulation"]
    roster_count = 0
    batch = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        rid = _clean_str(row[0])
        if not rid:
            continue
        batch.append(row)
        if len(batch) >= 500:
            _flush_roster(db, batch, counts, rule_version_id=rv.id)
            roster_count += len(batch)
            batch = []
    if batch:
        _flush_roster(db, batch, counts, rule_version_id=rv.id)
        roster_count += len(batch)
    counts["roster_total"] = roster_count

    # 14. Backfill rule_version_id on existing roster if missing
    if rv:
        updated = db.query(RosterAssignment).filter(
            RosterAssignment.tenant_id == TENANT_ID,
            RosterAssignment.rule_version_id == None,
        ).update({RosterAssignment.rule_version_id: rv.id})
        if updated:
            counts["roster_backfill_rv"] = updated

    db.commit()
    return counts


def _flush_roster(db: Session, rows: list, counts: dict, rule_version_id: str | None = None):
    """Insert a batch of roster rows."""
    for row in rows:
        rid = _clean_str(row[0])
        if not rid:
            continue
        if db.query(RosterAssignment).filter(RosterAssignment.id == rid).first():
            continue
        op_date = _parse_date(row[1])
        if not op_date:
            continue
        work_status_str = _clean_str(row[6]) or "WORK"
        site_status_str = _clean_str(row[5]) or "ONSITE"
        db.add(RosterAssignment(
            id=rid,
            tenant_id=TENANT_ID,
            roster_code=rid,
            operating_date=op_date,
            employee_id=_clean_str(row[2]) or rid,
            crew_id=_clean_str(row[3]),
            site_cycle_day=_clean_int(row[4]),
            site_status=SiteStatusEnum(site_status_str),
            work_status=WorkStatus(work_status_str),
            shift_id=_clean_str(row[7]),
            site_id=_clean_str(row[8]),
            planned_equipment_id=_clean_str(row[9]),
            rule_version_id=rule_version_id,
            effective_rule_version=_clean_str(row[10]),
            validation_status=ValidationStatus(_clean_str(row[11]) or "VALID"),
        ))
    db.flush()

    # 10. Tenant Capabilities (RosterPolicy)
    # Metro Mining: equipment assignment + competency validation enabled
    capability_policies = {
        "equipment_assignment_enabled": ("true", "boolean", "CONFIRMED",
            "Metro Mining tracks planned vs actual equipment assignment"),
        "competency_validation_enabled": ("true", "boolean", "CONFIRMED",
            "Metro Mining validates operator competency for equipment types"),
    }
    for key, (val, dtype, status, notes) in capability_policies.items():
        if not db.query(RosterPolicy).filter(
            RosterPolicy.tenant_id == TENANT_ID,
            RosterPolicy.policy_key == key,
        ).first():
            db.add(RosterPolicy(
                id=f"rp-cap-{key}",
                tenant_id=TENANT_ID,
                policy_key=key,
                policy_value=val,
                data_type=dtype,
                confirmation_status=status,
                notes=notes,
            ))
    db.flush()


def main():
    settings = get_settings()
    engine = create_engine(settings.database_url)
    Base.metadata.create_all(engine)
    with Session(engine) as db:
        counts = seed(db)
    print("\n=== Metro Mining Seed Complete ===")
    for k, v in counts.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
