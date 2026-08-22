"""
report_service.py — MM-M4D Reports & Export Service.

READ-ONLY reporting layer over existing M0–M4C data.
No new business rules. No new tables. Pure query + export.

Three reports:
  A) Shift Attendance Report
  B) Exception & Decision Report
  C) Roster vs Actual Equipment Report

Exports: CSV (UTF-8 BOM, formula-injection safe), XLSX (openpyxl).
Row limit: 1000 per export.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field, asdict
from datetime import date, datetime, time, timezone, timedelta
from typing import Any

from sqlalchemy import and_, or_, func, case, select
from sqlalchemy.orm import Session, aliased

from app.models import (
    RosterAssignment, Worker, EmployeeMeta, Role, Crew, Equipment,
    ShiftTemplate, Site, Tenant, RuleVersion,
    CanonicalAttendanceEvent, CanonicalEventType, CanonicalProcessingStatus,
    CheckpointValidationResult, CheckpointValidationStatus,
    MissingCheckpointResult,
    ExceptionCase, ExceptionStatus, ExceptionSeverity,
    ExceptionDecision, DecisionStatus, DecisionType,
    ExceptionAction, ExceptionActionType,
    ExceptionEvidence,
    EquipmentAssignmentActual, ActualAssignmentStatus,
    EquipmentComparisonResult, ComparisonResult,
    EquipmentDiscrepancy, DiscrepancyType, DiscrepancyStatus,
    WorkStatus, SiteStatusEnum,
    User,
)

# ── Constants ─────────────────────────────────────────────────
MAX_EXPORT_ROWS = 1000

# Checkpoint type strings used in checkpoint_validation_results
# (matches checkpoint_policies seed: BRIEFING_IN, EQUIPMENT_IN, WORK_START,
#  BREAK_OUT, BREAK_IN, HANDOVER, SHIFT_OUT)
_CP_BRIEFING_IN = "BRIEFING_IN"
_CP_EQUIPMENT_IN = "EQUIPMENT_IN"
_CP_WORK_START = "WORK_START"
_CP_BREAK_OUT = "BREAK_OUT"
_CP_BREAK_IN = "BREAK_IN"
_CP_HANDOVER = "HANDOVER"
_CP_SHIFT_OUT = "SHIFT_OUT"

# Non-work statuses that should have None checkpoint fields
_NON_WORK_STATUSES = {WorkStatus.REST, WorkStatus.OFFSITE, WorkStatus.LEAVE, WorkStatus.SICK}


# ── Dataclasses ───────────────────────────────────────────────

@dataclass
class ReportFilter:
    """Common filter for all reports."""
    tenant_id: str
    operating_date: date | None = None
    date_from: date | None = None
    date_to: date | None = None
    site_id: str | None = None
    shift_id: str | None = None
    crew_id: str | None = None
    role_id: str | None = None
    employee_id: str | None = None
    equipment_id: str | None = None
    work_status: str | None = None
    exception_type: str | None = None
    severity: str | None = None
    exception_status: str | None = None
    decision_status: str | None = None


@dataclass
class ReportMetadata:
    """Metadata attached to every export."""
    tenant_name: str
    tenant_code: str
    site_name: str | None
    operating_date: date | None
    date_from: date | None
    date_to: date | None
    shift_name: str | None
    timezone: str
    generated_at: datetime
    report_type: str
    applied_filters: dict
    row_count: int


@dataclass
class ShiftAttendanceRow:
    """Report A row."""
    operating_date: date
    site_name: str
    shift_name: str
    employee_no: str
    employee_name: str
    role_name: str
    crew_name: str
    work_status: str
    planned_equipment: str | None
    briefing_in: datetime | None
    equipment_in: datetime | None
    work_start: datetime | None
    break_out: datetime | None
    break_in: datetime | None
    handover: datetime | None
    shift_out: datetime | None
    operational_state: str
    attendance_exception_count: int
    exception_types: str
    rule_version: str | None


@dataclass
class ExceptionDecisionRow:
    """Report B row."""
    exception_id: str
    operating_date: date
    shift_name: str
    employee_no: str
    employee_name: str
    role_name: str
    crew_name: str
    equipment_code: str | None
    exception_type: str
    severity: str
    detected_at: datetime
    status: str
    current_owner: str | None
    acknowledged_at: datetime | None
    resolved_at: datetime | None
    waived_at: datetime | None
    decision_type: str | None
    decision_status: str | None
    requested_by: str | None
    decided_by: str | None
    decided_at: datetime | None
    decision_reason: str | None
    rule_version: str | None
    source_type: str


@dataclass
class RosterVsActualRow:
    """Report C row."""
    operating_date: date
    shift_name: str
    site_name: str
    employee_no: str
    employee_name: str
    role_name: str
    crew_name: str
    planned_equipment: str
    actual_equipment: str | None
    actual_operator: str | None
    actual_start: datetime | None
    actual_end: datetime | None
    comparison_result: str | None
    discrepancy_type: str | None
    exception_status: str | None
    decision_type: str | None
    decision_status: str | None
    decision_reason: str | None


# ── Column definitions for exports ───────────────────────────

SHIFT_ATTENDANCE_COLUMNS = [
    "operating_date", "site_name", "shift_name", "employee_no",
    "employee_name", "role_name", "crew_name", "work_status",
    "planned_equipment", "briefing_in", "equipment_in", "work_start",
    "break_out", "break_in", "handover", "shift_out",
    "operational_state", "attendance_exception_count", "exception_types",
    "rule_version",
]

SHIFT_ATTENDANCE_HEADERS = {
    "operating_date": "Operating Date",
    "site_name": "Site",
    "shift_name": "Shift",
    "employee_no": "Employee No",
    "employee_name": "Employee Name",
    "role_name": "Role",
    "crew_name": "Crew",
    "work_status": "Work Status",
    "planned_equipment": "Planned Equipment",
    "briefing_in": "Briefing In",
    "equipment_in": "Equipment In",
    "work_start": "Work Start",
    "break_out": "Break Out",
    "break_in": "Break In",
    "handover": "Handover",
    "shift_out": "Shift Out",
    "operational_state": "Operational State",
    "attendance_exception_count": "Exception Count",
    "exception_types": "Exception Types",
    "rule_version": "Rule Version",
}

EXCEPTION_DECISION_COLUMNS = [
    "exception_id", "operating_date", "shift_name", "employee_no",
    "employee_name", "role_name", "crew_name", "equipment_code",
    "exception_type", "severity", "detected_at", "status",
    "current_owner", "acknowledged_at", "resolved_at", "waived_at",
    "decision_type", "decision_status", "requested_by", "decided_by",
    "decided_at", "decision_reason", "rule_version", "source_type",
]

EXCEPTION_DECISION_HEADERS = {
    "exception_id": "Exception ID",
    "operating_date": "Operating Date",
    "shift_name": "Shift",
    "employee_no": "Employee No",
    "employee_name": "Employee Name",
    "role_name": "Role",
    "crew_name": "Crew",
    "equipment_code": "Equipment",
    "exception_type": "Exception Type",
    "severity": "Severity",
    "detected_at": "Detected At",
    "status": "Status",
    "current_owner": "Current Owner",
    "acknowledged_at": "Acknowledged At",
    "resolved_at": "Resolved At",
    "waived_at": "Waived At",
    "decision_type": "Decision Type",
    "decision_status": "Decision Status",
    "requested_by": "Requested By",
    "decided_by": "Decided By",
    "decided_at": "Decided At",
    "decision_reason": "Decision Reason",
    "rule_version": "Rule Version",
    "source_type": "Source Type",
}

ROSTER_VS_ACTUAL_COLUMNS = [
    "operating_date", "shift_name", "site_name", "employee_no",
    "employee_name", "role_name", "crew_name", "planned_equipment",
    "actual_equipment", "actual_operator", "actual_start", "actual_end",
    "comparison_result", "discrepancy_type", "exception_status",
    "decision_type", "decision_status", "decision_reason",
]

ROSTER_VS_ACTUAL_HEADERS = {
    "operating_date": "Operating Date",
    "shift_name": "Shift",
    "site_name": "Site",
    "employee_no": "Employee No",
    "employee_name": "Employee Name",
    "role_name": "Role",
    "crew_name": "Crew",
    "planned_equipment": "Planned Equipment",
    "actual_equipment": "Actual Equipment",
    "actual_operator": "Actual Operator",
    "actual_start": "Actual Start",
    "actual_end": "Actual End",
    "comparison_result": "Comparison Result",
    "discrepancy_type": "Discrepancy Type",
    "exception_status": "Exception Status",
    "decision_type": "Decision Type",
    "decision_status": "Decision Status",
    "decision_reason": "Decision Reason",
}


# ── Helpers ───────────────────────────────────────────────────

def _sanitize_cell(value: Any) -> Any:
    """Prefix dangerous formula characters with single quote."""
    if isinstance(value, str) and value and value[0] in ('=', '+', '-', '@', '\t', '\r'):
        return "'" + value
    return value


def _str_or_none(val: Any) -> str | None:
    """Convert to string or return None."""
    if val is None:
        return None
    return str(val)


def _enum_val(val: Any) -> str:
    """Extract .value from enum or return str."""
    if val is None:
        return ""
    return val.value if hasattr(val, "value") else str(val)


def _resolve_operational_state(
    work_status: WorkStatus,
    checkpoint_map: dict[str, datetime | None],
) -> str:
    """Derive operational state from checkpoint timestamps.

    For non-WORK statuses, returns the work_status directly.
    For WORK, derives from the latest checkpoint received.
    """
    if work_status != WorkStatus.WORK:
        return _enum_val(work_status)

    # Check in reverse chronological order of a typical shift
    if checkpoint_map.get(_CP_SHIFT_OUT):
        return "OFF_DUTY"
    if checkpoint_map.get(_CP_HANDOVER):
        return "HANDOVER"
    if checkpoint_map.get(_CP_BREAK_OUT) and not checkpoint_map.get(_CP_BREAK_IN):
        return "ON_BREAK"
    if checkpoint_map.get(_CP_WORK_START) or checkpoint_map.get(_CP_BRIEFING_IN):
        return "WORK"

    return "ABSENT"


def _build_date_filter(
    query, model, f: ReportFilter,
):
    """Apply date range / single date filter to a query."""
    if f.operating_date:
        query = query.where(model.operating_date == f.operating_date)
    else:
        if f.date_from:
            query = query.where(model.operating_date >= f.date_from)
        if f.date_to:
            query = query.where(model.operating_date <= f.date_to)
    return query


def _build_common_filters(
    query, ra: RosterAssignment, f: ReportFilter,
):
    """Apply common roster-level filters."""
    query = query.where(ra.tenant_id == f.tenant_id)
    query = _build_date_filter(query, ra, f)
    if f.site_id:
        query = query.where(ra.site_id == f.site_id)
    if f.shift_id:
        query = query.where(ra.shift_id == f.shift_id)
    if f.crew_id:
        query = query.where(ra.crew_id == f.crew_id)
    if f.employee_id:
        query = query.where(ra.employee_id == f.employee_id)
    if f.work_status:
        query = query.where(ra.work_status == f.work_status)
    return query


def _user_display_name_map(db: Session, user_ids: set[str]) -> dict[str, str]:
    """Batch-load user display names."""
    if not user_ids:
        return {}
    users = db.query(User.id, User.display_name).filter(User.id.in_(user_ids)).all()
    return {u.id: u.display_name for u in users}


# ── Report A: Shift Attendance ────────────────────────────────

def query_shift_attendance(db: Session, f: ReportFilter) -> list[ShiftAttendanceRow]:
    """Query shift attendance report.

    Joins roster_assignments → workers, employee_meta, roles, crews,
    shift_templates, sites. LEFT JOINs for equipment, checkpoints,
    canonical events, and exceptions.
    """
    ra = RosterAssignment
    w = Worker
    em = EmployeeMeta
    r = Role
    c = Crew
    st = ShiftTemplate
    s = Site
    eq = Equipment
    cvr = CheckpointValidationResult
    cae = CanonicalAttendanceEvent
    ec = ExceptionCase
    rv = RuleVersion

    # Base query: roster_assignments with joins
    query = (
        db.query(
            ra.operating_date,
            s.site_name,
            st.shift_name,
            em.employee_no,
            w.name.label("employee_name"),
            r.role_name,
            c.crew_name,
            ra.work_status,
            eq.equipment_code.label("planned_equipment"),
            ra.rule_version_id,
            rv.version_label.label("rule_version_label"),
            ra.employee_id,
            ra.shift_id,
            ra.id.label("roster_id"),
        )
        .join(w, w.id == ra.employee_id)
        .outerjoin(em, and_(
            em.worker_id == w.id,
            em.tenant_id == ra.tenant_id,
        ))
        .outerjoin(r, and_(
            r.id == em.role_id,
            r.tenant_id == ra.tenant_id,
        ))
        .outerjoin(c, and_(
            c.id == ra.crew_id,
            c.tenant_id == ra.tenant_id,
        ))
        .outerjoin(st, and_(
            st.id == ra.shift_id,
            st.tenant_id == ra.tenant_id,
        ))
        .outerjoin(s, and_(
            s.id == ra.site_id,
            s.tenant_id == ra.tenant_id,
        ))
        .outerjoin(eq, and_(
            eq.id == ra.planned_equipment_id,
            eq.tenant_id == ra.tenant_id,
        ))
        .outerjoin(rv, and_(
            rv.id == ra.rule_version_id,
            rv.tenant_id == ra.tenant_id,
        ))
    )

    query = _build_common_filters(query, ra, f)
    if f.role_id:
        query = query.where(em.role_id == f.role_id)

    # Deterministic ordering: operating_date → shift → crew → employee
    query = query.order_by(
        ra.operating_date,
        st.shift_name.nullsfirst(),
        c.crew_name.nullsfirst(),
        em.employee_no.nullsfirst(),
    )

    query = query.limit(MAX_EXPORT_ROWS)
    rows = query.all()

    if not rows:
        return []

    # Collect roster IDs for batch-loading checkpoints and exceptions
    roster_ids = [r.roster_id for r in rows]
    employee_ids = [r.employee_id for r in rows]
    operating_dates = list({r.operating_date for r in rows})
    shift_ids = list({r.shift_id for r in rows if r.shift_id})

    # Batch-load checkpoint validation results
    # Key: (employee_id, operating_date, shift_id, checkpoint_type) → detected_timestamp
    checkpoint_data: dict[tuple, datetime] = {}
    if roster_ids:
        cp_rows = (
            db.query(
                cvr.employee_id,
                cvr.operating_date,
                cvr.shift_id,
                cvr.checkpoint_type,
                cvr.detected_timestamp,
            )
            .filter(
                cvr.tenant_id == f.tenant_id,
                cvr.employee_id.in_(employee_ids),
                cvr.operating_date.in_(operating_dates),
            )
            .all()
        )
        for cp in cp_rows:
            key = (cp.employee_id, cp.operating_date, cp.shift_id, cp.checkpoint_type)
            # Keep the latest timestamp per checkpoint type
            if key not in checkpoint_data or cp.detected_timestamp > checkpoint_data[key]:
                checkpoint_data[key] = cp.detected_timestamp

    # Batch-load exception counts per (employee_id, operating_date, shift_id)
    exception_counts: dict[tuple, tuple[int, str]] = {}
    if employee_ids and operating_dates:
        exc_rows = (
            db.query(
                ec.employee_id,
                ec.operating_date,
                ec.shift_id,
                func.count(ec.id).label("cnt"),
                func.group_concat(ec.exception_type, ",").label("types"),
            )
            .filter(
                ec.tenant_id == f.tenant_id,
                ec.employee_id.in_(employee_ids),
                ec.operating_date.in_(operating_dates),
            )
            .group_by(ec.employee_id, ec.operating_date, ec.shift_id)
            .all()
        )
        for ex in exc_rows:
            key = (ex.employee_id, ex.operating_date, ex.shift_id)
            exception_counts[key] = (ex.cnt, ex.types or "")

    # Build result rows
    result: list[ShiftAttendanceRow] = []
    for row in rows:
        ws = row.work_status
        is_non_work = ws in _NON_WORK_STATUSES

        # For non-WORK statuses, checkpoint fields are None
        if is_non_work:
            cp_map: dict[str, datetime | None] = {}
            briefing_in = None
            equipment_in = None
            work_start = None
            break_out_ts = None
            break_in_ts = None
            handover_ts = None
            shift_out_ts = None
        else:
            # Look up checkpoints by (employee_id, operating_date, shift_id)
            def _get_cp(cp_type: str) -> datetime | None:
                key = (row.employee_id, row.operating_date, row.shift_id, cp_type)
                return checkpoint_data.get(key)

            briefing_in = _get_cp(_CP_BRIEFING_IN)
            equipment_in = _get_cp(_CP_EQUIPMENT_IN)
            work_start = _get_cp(_CP_WORK_START)
            break_out_ts = _get_cp(_CP_BREAK_OUT)
            break_in_ts = _get_cp(_CP_BREAK_IN)
            handover_ts = _get_cp(_CP_HANDOVER)
            shift_out_ts = _get_cp(_CP_SHIFT_OUT)

            cp_map = {
                _CP_BRIEFING_IN: briefing_in,
                _CP_EQUIPMENT_IN: equipment_in,
                _CP_WORK_START: work_start,
                _CP_BREAK_OUT: break_out_ts,
                _CP_BREAK_IN: break_in_ts,
                _CP_HANDOVER: handover_ts,
                _CP_SHIFT_OUT: shift_out_ts,
            }

        operational_state = _resolve_operational_state(ws, cp_map)

        # Exception data
        exc_key = (row.employee_id, row.operating_date, row.shift_id)
        exc_count, exc_types = exception_counts.get(exc_key, (0, ""))

        result.append(ShiftAttendanceRow(
            operating_date=row.operating_date,
            site_name=row.site_name or "",
            shift_name=row.shift_name or "",
            employee_no=row.employee_no or "",
            employee_name=row.employee_name or "",
            role_name=row.role_name or "",
            crew_name=row.crew_name or "",
            work_status=_enum_val(ws),
            planned_equipment=row.planned_equipment,
            briefing_in=briefing_in,
            equipment_in=equipment_in,
            work_start=work_start,
            break_out=break_out_ts,
            break_in=break_in_ts,
            handover=handover_ts,
            shift_out=shift_out_ts,
            operational_state=operational_state,
            attendance_exception_count=exc_count,
            exception_types=exc_types,
            rule_version=row.rule_version_label,
        ))

    return result


# ── Report B: Exception & Decision ────────────────────────────

def query_exception_decisions(db: Session, f: ReportFilter) -> list[ExceptionDecisionRow]:
    """Query exception & decision report.

    Joins exception_cases → workers, employee_meta, roles, crews,
    equipments, rule_versions. LEFT JOINs for latest decision and
    action timestamps.
    """
    ec = ExceptionCase
    w = Worker
    em = EmployeeMeta
    r = Role
    c = Crew
    eq = Equipment
    rv = RuleVersion
    st = ShiftTemplate
    ed = ExceptionDecision
    ea = ExceptionAction

    # Subquery: latest decision per exception
    latest_dec_sq = (
        db.query(
            ed.exception_id,
            func.max(ed.created_at).label("max_created"),
        )
        .filter(ed.tenant_id == f.tenant_id)
        .group_by(ed.exception_id)
        .subquery()
    )

    # Subquery: acknowledged_at from exception_actions
    ack_sq = (
        db.query(
            ea.exception_id,
            func.min(ea.action_timestamp).label("acknowledged_at"),
        )
        .filter(
            ea.tenant_id == f.tenant_id,
            ea.new_status == ExceptionStatus.ACKNOWLEDGED,
        )
        .group_by(ea.exception_id)
        .subquery()
    )

    # Subquery: resolved_at from exception_actions
    res_sq = (
        db.query(
            ea.exception_id,
            func.min(ea.action_timestamp).label("resolved_at"),
        )
        .filter(
            ea.tenant_id == f.tenant_id,
            ea.new_status == ExceptionStatus.RESOLVED,
        )
        .group_by(ea.exception_id)
        .subquery()
    )

    # Subquery: waived_at from exception_actions
    wav_sq = (
        db.query(
            ea.exception_id,
            func.min(ea.action_timestamp).label("waived_at"),
        )
        .filter(
            ea.tenant_id == f.tenant_id,
            ea.new_status == ExceptionStatus.WAIVED,
        )
        .group_by(ea.exception_id)
        .subquery()
    )

    # Requested-by user names subquery
    req_user_sq = (
        db.query(User.id, User.display_name)
        .subquery()
    )

    query = (
        db.query(
            ec.id.label("exception_id"),
            ec.operating_date,
            st.shift_name,
            em.employee_no,
            w.name.label("employee_name"),
            r.role_name,
            c.crew_name,
            eq.equipment_code,
            ec.exception_type,
            ec.severity,
            ec.detected_at,
            ec.status,
            ec.current_owner_id,
            ec.source_type,
            ec.rule_version_id,
            rv.version_label.label("rule_version_label"),
            # Decision fields
            ed.decision_type.label("dec_decision_type"),
            ed.status.label("dec_status"),
            ed.requested_by.label("dec_requested_by"),
            ed.decided_by.label("dec_decided_by"),
            ed.decided_at.label("dec_decided_at"),
            ed.reason_text.label("dec_reason_text"),
            # Action timestamps
            ack_sq.c.acknowledged_at,
            res_sq.c.resolved_at,
            wav_sq.c.waived_at,
        )
        .join(w, w.id == ec.employee_id)
        .outerjoin(em, and_(
            em.worker_id == w.id,
            em.tenant_id == ec.tenant_id,
        ))
        .outerjoin(r, and_(
            r.id == em.role_id,
            r.tenant_id == ec.tenant_id,
        ))
        .outerjoin(c, and_(
            c.id == em.crew_id,
            c.tenant_id == ec.tenant_id,
        ))
        .outerjoin(eq, and_(
            eq.id == ec.equipment_id,
            eq.tenant_id == ec.tenant_id,
        ))
        .outerjoin(st, and_(
            st.id == ec.shift_id,
            st.tenant_id == ec.tenant_id,
        ))
        .outerjoin(rv, and_(
            rv.id == ec.rule_version_id,
            rv.tenant_id == ec.tenant_id,
        ))
        .outerjoin(latest_dec_sq, and_(
            latest_dec_sq.c.exception_id == ec.id,
        ))
        .outerjoin(ed, and_(
            ed.exception_id == ec.id,
            ed.created_at == latest_dec_sq.c.max_created,
        ))
        .outerjoin(ack_sq, ack_sq.c.exception_id == ec.id)
        .outerjoin(res_sq, res_sq.c.exception_id == ec.id)
        .outerjoin(wav_sq, wav_sq.c.exception_id == ec.id)
    )

    # Tenant filter first
    query = query.where(ec.tenant_id == f.tenant_id)
    query = _build_date_filter(query, ec, f)

    if f.site_id:
        query = query.where(ec.site_id == f.site_id)
    if f.shift_id:
        query = query.where(ec.shift_id == f.shift_id)
    if f.employee_id:
        query = query.where(ec.employee_id == f.employee_id)
    if f.equipment_id:
        query = query.where(ec.equipment_id == f.equipment_id)
    if f.exception_type:
        query = query.where(ec.exception_type == f.exception_type)
    if f.severity:
        query = query.where(ec.severity == f.severity)
    if f.exception_status:
        query = query.where(ec.status == f.exception_status)
    if f.decision_status:
        query = query.where(ed.status == f.decision_status)

    # Deterministic ordering
    query = query.order_by(
        ec.operating_date,
        st.shift_name.nullsfirst(),
        c.crew_name.nullsfirst(),
        em.employee_no.nullsfirst(),
        ec.detected_at,
    )

    query = query.limit(MAX_EXPORT_ROWS)
    rows = query.all()

    if not rows:
        return []

    # Batch-load user display names for owners, requesters, deciders
    user_ids: set[str] = set()
    for row in rows:
        if row.current_owner_id:
            user_ids.add(row.current_owner_id)
        if row.dec_requested_by:
            user_ids.add(row.dec_requested_by)
        if row.dec_decided_by:
            user_ids.add(row.dec_decided_by)
    user_names = _user_display_name_map(db, user_ids)

    result: list[ExceptionDecisionRow] = []
    for row in rows:
        result.append(ExceptionDecisionRow(
            exception_id=str(row.exception_id),
            operating_date=row.operating_date,
            shift_name=row.shift_name or "",
            employee_no=row.employee_no or "",
            employee_name=row.employee_name or "",
            role_name=row.role_name or "",
            crew_name=row.crew_name or "",
            equipment_code=row.equipment_code,
            exception_type=row.exception_type or "",
            severity=_enum_val(row.severity),
            detected_at=row.detected_at,
            status=_enum_val(row.status),
            current_owner=user_names.get(row.current_owner_id) if row.current_owner_id else None,
            acknowledged_at=row.acknowledged_at,
            resolved_at=row.resolved_at,
            waived_at=row.waived_at,
            decision_type=_enum_val(row.dec_decision_type) if row.dec_decision_type else None,
            decision_status=_enum_val(row.dec_status) if row.dec_status else None,
            requested_by=user_names.get(row.dec_requested_by) if row.dec_requested_by else None,
            decided_by=user_names.get(row.dec_decided_by) if row.dec_decided_by else None,
            decided_at=row.dec_decided_at,
            decision_reason=row.dec_reason_text,
            rule_version=row.rule_version_label,
            source_type=row.source_type or "",
        ))

    return result


# ── Report C: Roster vs Actual Equipment ──────────────────────

def query_roster_vs_actual(db: Session, f: ReportFilter) -> list[RosterVsActualRow]:
    """Query roster vs actual equipment report.

    Joins roster_assignments → workers, equipments (planned), sites,
    shift_templates. LEFT JOINs for actual assignments, comparison
    results, discrepancies, and substitution decisions.
    """
    ra = RosterAssignment
    w = Worker
    em = EmployeeMeta
    r = Role
    c = Crew
    st = ShiftTemplate
    s = Site
    eq_plan = Equipment
    eq_act = aliased(Equipment, name="eq_actual")
    eaa = EquipmentAssignmentActual
    ecr = EquipmentComparisonResult
    edis = EquipmentDiscrepancy
    ed = ExceptionDecision
    ec = ExceptionCase

    # Subquery: latest decision per exception (for substitution decisions)
    latest_dec_sq = (
        db.query(
            ed.exception_id,
            func.max(ed.created_at).label("max_created"),
        )
        .filter(ed.tenant_id == f.tenant_id)
        .group_by(ed.exception_id)
        .subquery()
    )

    # Subquery: discrepancy per actual_assignment
    disc_sq = (
        db.query(
            edis.id.label("disc_id"),
            edis.actual_assignment_id,
            edis.discrepancy_type,
            edis.status.label("disc_status"),
        )
        .filter(edis.tenant_id == f.tenant_id)
        .subquery()
    )

    # Subquery: exception linked to discrepancy
    exc_for_disc_sq = (
        db.query(
            ec.id.label("exc_id"),
            ec.source_id.label("discrepancy_id"),
            ec.status.label("exc_status"),
        )
        .filter(
            ec.tenant_id == f.tenant_id,
            ec.source_type == "EQUIPMENT_DISCREPANCY",
        )
        .subquery()
    )

    # Actual operator name (worker who actually operated)
    actual_op_w = Worker.__table__.alias("actual_op_w")

    query = (
        db.query(
            ra.operating_date,
            st.shift_name,
            s.site_name,
            em.employee_no,
            w.name.label("employee_name"),
            r.role_name,
            c.crew_name,
            eq_plan.equipment_code.label("planned_equipment"),
            # Actual assignment fields
            eaa.id.label("actual_id"),
            eq_act.equipment_code.label("actual_equipment"),
            eaa.started_at.label("actual_start"),
            eaa.ended_at.label("actual_end"),
            eaa.employee_id.label("actual_employee_id"),
            # Comparison
            ecr.comparison_result,
            ecr.actual_worker_id,
            # Discrepancy
            disc_sq.c.discrepancy_type,
            disc_sq.c.disc_status,
            # Exception status from discrepancy
            exc_for_disc_sq.c.exc_status,
            # Decision (from exception linked to discrepancy)
            ed.decision_type.label("dec_decision_type"),
            ed.status.label("dec_status"),
            ed.reason_text.label("dec_reason_text"),
        )
        .join(w, w.id == ra.employee_id)
        .outerjoin(em, and_(
            em.worker_id == w.id,
            em.tenant_id == ra.tenant_id,
        ))
        .outerjoin(r, and_(
            r.id == em.role_id,
            r.tenant_id == ra.tenant_id,
        ))
        .outerjoin(c, and_(
            c.id == em.crew_id,
            c.tenant_id == ra.tenant_id,
        ))
        .outerjoin(st, and_(
            st.id == ra.shift_id,
            st.tenant_id == ra.tenant_id,
        ))
        .outerjoin(s, and_(
            s.id == ra.site_id,
            s.tenant_id == ra.tenant_id,
        ))
        .outerjoin(eq_plan, and_(
            eq_plan.id == ra.planned_equipment_id,
            eq_plan.tenant_id == ra.tenant_id,
        ))
        # LEFT JOIN actual assignments (multiple per shift)
        .outerjoin(eaa, and_(
            eaa.roster_id == ra.id,
            eaa.tenant_id == ra.tenant_id,
        ))
        # LEFT JOIN actual equipment
        .outerjoin(eq_act, and_(
            eq_act.id == eaa.equipment_id,
            eq_act.tenant_id == ra.tenant_id,
        ))
        # LEFT JOIN comparison results
        .outerjoin(ecr, and_(
            ecr.actual_assignment_id == eaa.id,
            ecr.tenant_id == ra.tenant_id,
        ))
        # LEFT JOIN discrepancy
        .outerjoin(disc_sq, disc_sq.c.actual_assignment_id == eaa.id)
        # LEFT JOIN exception for discrepancy
        .outerjoin(exc_for_disc_sq, and_(
            exc_for_disc_sq.c.discrepancy_id == disc_sq.c.disc_id,
        ))
        # LEFT JOIN latest decision for that exception
        .outerjoin(latest_dec_sq, and_(
            latest_dec_sq.c.exception_id == exc_for_disc_sq.c.exc_id,
        ))
        .outerjoin(ed, and_(
            ed.exception_id == latest_dec_sq.c.exception_id,
            ed.created_at == latest_dec_sq.c.max_created,
        ))
    )

    # Tenant filter first
    query = query.where(ra.tenant_id == f.tenant_id)
    query = _build_date_filter(query, ra, f)

    # Only WORK status employees (REST/OFFSITE have no equipment assignments)
    query = query.where(ra.work_status == WorkStatus.WORK)

    if f.site_id:
        query = query.where(ra.site_id == f.site_id)
    if f.shift_id:
        query = query.where(ra.shift_id == f.shift_id)
    if f.crew_id:
        query = query.where(ra.crew_id == f.crew_id)
    if f.employee_id:
        query = query.where(ra.employee_id == f.employee_id)
    if f.equipment_id:
        query = query.where(
            or_(
                ra.planned_equipment_id == f.equipment_id,
                eaa.equipment_id == f.equipment_id,
            )
        )

    # Deterministic ordering
    query = query.order_by(
        ra.operating_date,
        st.shift_name.nullsfirst(),
        c.crew_name.nullsfirst(),
        em.employee_no.nullsfirst(),
        eaa.started_at.nullsfirst(),
    )

    query = query.limit(MAX_EXPORT_ROWS)
    rows = query.all()

    if not rows:
        return []

    # Batch-load actual operator names (for operator substitution)
    actual_worker_ids = {r.actual_worker_id for r in rows if r.actual_worker_id}
    actual_op_names: dict[str, str] = {}
    if actual_worker_ids:
        op_workers = db.query(Worker.id, Worker.name).filter(
            Worker.id.in_(actual_worker_ids)
        ).all()
        actual_op_names = {w.id: w.name for w in op_workers}

    result: list[RosterVsActualRow] = []
    for row in rows:
        # Determine actual operator name
        actual_operator = None
        if row.actual_worker_id and row.actual_worker_id != row.actual_employee_id:
            # Different worker → operator substitution
            actual_operator = actual_op_names.get(row.actual_worker_id, row.actual_worker_id)

        result.append(RosterVsActualRow(
            operating_date=row.operating_date,
            shift_name=row.shift_name or "",
            site_name=row.site_name or "",
            employee_no=row.employee_no or "",
            employee_name=row.employee_name or "",
            role_name=row.role_name or "",
            crew_name=row.crew_name or "",
            planned_equipment=row.planned_equipment or "",
            actual_equipment=row.actual_equipment,
            actual_operator=actual_operator,
            actual_start=row.actual_start,
            actual_end=row.actual_end,
            comparison_result=_enum_val(row.comparison_result) if row.comparison_result else None,
            discrepancy_type=_enum_val(row.discrepancy_type) if row.discrepancy_type else None,
            exception_status=_enum_val(row.exc_status) if row.exc_status else None,
            decision_type=_enum_val(row.dec_decision_type) if row.dec_decision_type else None,
            decision_status=_enum_val(row.dec_status) if row.dec_status else None,
            decision_reason=row.dec_reason_text,
        ))

    return result


# ── Export: CSV ───────────────────────────────────────────────

def export_csv(rows: list[dict], columns: list[str]) -> str:
    """Export rows to CSV string. UTF-8 with BOM. Formula injection safe."""
    output = io.StringIO()
    # UTF-8 BOM
    output.write('\ufeff')
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction='ignore')
    writer.writeheader()
    for row in rows:
        sanitized = {k: _sanitize_cell(v) for k, v in row.items() if k in columns}
        # Format datetimes
        for k, v in sanitized.items():
            if isinstance(v, datetime):
                sanitized[k] = v.strftime("%Y-%m-%d %H:%M:%S")
            elif isinstance(v, date):
                sanitized[k] = v.strftime("%Y-%m-%d")
            elif v is None:
                sanitized[k] = ""
        writer.writerow(sanitized)
    return output.getvalue()


# ── Export: XLSX ──────────────────────────────────────────────

def export_xlsx(
    rows: list[dict],
    columns: list[str],
    sheet_name: str,
    metadata,
    filename_hint: str,
) -> bytes:
    """Export to XLSX bytes. Uses openpyxl.

    Requirements:
    - Proper column headers (human-readable)
    - Auto-width columns (approximate)
    - Date/time formatting (YYYY-MM-DD HH:MM:SS)
    - Metadata sheet ("Report_Info")
    - Formula injection protection
    - Deterministic sheet names
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Data sheet ──
    ws = wb.active
    # Sanitize sheet name (max 31 chars, no special chars)
    safe_sheet_name = re.sub(r'[\\/*?\[\]:]', '_', sheet_name)[:31]
    ws.title = safe_sheet_name

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Determine headers map — use the appropriate headers dict
    headers_map = _get_headers_for_columns(columns)

    # Write headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=headers_map.get(col_name, col_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Track column widths
    col_widths = {col_idx: len(headers_map.get(col_name, col_name))
                  for col_idx, col_name in enumerate(columns, 1)}

    # Date/time format
    dt_format = "YYYY-MM-DD HH:MM:SS"
    d_format = "YYYY-MM-DD"

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name)

            # Formula injection protection
            value = _sanitize_cell(value)

            cell = ws.cell(row=row_idx, column=col_idx)

            if value is None:
                cell.value = ""
            elif isinstance(value, datetime):
                cell.value = value
                cell.number_format = dt_format
            elif isinstance(value, date):
                cell.value = value
                cell.number_format = d_format
            else:
                cell.value = value

            cell.border = thin_border

            # Track width
            str_val = str(value) if value is not None else ""
            col_widths[col_idx] = max(col_widths.get(col_idx, 0), min(len(str_val) + 2, 50))

    # Apply column widths
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 8)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    # ── Normalize metadata to dict ──
    if hasattr(metadata, '__dataclass_fields__'):
        meta = asdict(metadata)
    elif isinstance(metadata, dict):
        meta = metadata
    else:
        meta = {}

    # ── Metadata sheet ──
    ws_meta = wb.create_sheet("Report_Info")
    meta_font = Font(bold=True)
    meta_entries = [
        ("Tenant", meta.get("tenant_name", "")),
        ("Tenant Code", meta.get("tenant_code", "")),
        ("Site", meta.get("site_name", "All")),
        ("Operating Date", str(meta.get("operating_date", ""))),
        ("Date From", str(meta.get("date_from", ""))),
        ("Date To", str(meta.get("date_to", ""))),
        ("Shift", meta.get("shift_name", "All")),
        ("Timezone", meta.get("timezone", "")),
        ("Generated At", str(meta.get("generated_at", ""))),
        ("Report Type", meta.get("report_type", "")),
        ("Row Count", meta.get("row_count", 0)),
        ("Max Rows", MAX_EXPORT_ROWS),
    ]

    # Add applied filters
    applied = meta.get("applied_filters", {})

    ws_meta.cell(row=1, column=1, value="Field").font = meta_font
    ws_meta.cell(row=1, column=2, value="Value").font = meta_font
    ws_meta.column_dimensions["A"].width = 20
    ws_meta.column_dimensions["B"].width = 50

    for i, (key, val) in enumerate(meta_entries, 2):
        ws_meta.cell(row=i, column=1, value=key)
        ws_meta.cell(row=i, column=2, value=str(val) if val is not None else "")

    # Filters section
    filter_start = len(meta_entries) + 3
    ws_meta.cell(row=filter_start, column=1, value="Applied Filters").font = meta_font
    for i, (fk, fv) in enumerate(applied.items(), filter_start + 1):
        ws_meta.cell(row=i, column=1, value=fk)
        ws_meta.cell(row=i, column=2, value=str(fv) if fv is not None else "")

    # Save to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _get_headers_for_columns(columns: list[str]) -> dict[str, str]:
    """Get human-readable headers for a column list."""
    all_headers = {}
    all_headers.update(SHIFT_ATTENDANCE_HEADERS)
    all_headers.update(EXCEPTION_DECISION_HEADERS)
    all_headers.update(ROSTER_VS_ACTUAL_HEADERS)
    return {col: all_headers.get(col, col) for col in columns}


# ── Filename Generation ──────────────────────────────────────

def generate_filename(
    report_type: str,
    tenant_slug: str,
    operating_date: date | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    shift_id: str | None = None,
    fmt: str = "xlsx",
) -> str:
    """Generate a deterministic, human-readable filename.

    Examples:
        metro_shift_attendance_2026-09-05_DAY.xlsx
        metro_exceptions_2026-09-01_to_2026-09-07.xlsx
        metro_roster_vs_actual_2026-09-05_NIGHT.xlsx
    """
    parts = [tenant_slug, report_type]

    if operating_date:
        parts.append(operating_date.strftime("%Y-%m-%d"))
    elif date_from and date_to:
        parts.append(f"{date_from.strftime('%Y-%m-%d')}_to_{date_to.strftime('%Y-%m-%d')}")
    elif date_from:
        parts.append(f"from_{date_from.strftime('%Y-%m-%d')}")
    elif date_to:
        parts.append(f"to_{date_to.strftime('%Y-%m-%d')}")

    if shift_id:
        parts.append(shift_id)

    return "_".join(parts) + f".{fmt}"


# ── High-level export orchestration ──────────────────────────

def _build_applied_filters(f: ReportFilter) -> dict:
    """Extract non-None filter values for metadata."""
    result = {}
    d = asdict(f)
    for k, v in d.items():
        if k == "tenant_id":
            continue
        if v is not None:
            result[k] = str(v)
    return result


def _row_to_dict(row: Any) -> dict:
    """Convert a dataclass row to dict."""
    if hasattr(row, "__dataclass_fields__"):
        return asdict(row)
    return dict(row)


def export_shift_attendance_csv(db: Session, f: ReportFilter) -> tuple[str, str]:
    """Export Report A as CSV. Returns (csv_string, filename)."""
    rows = query_shift_attendance(db, f)
    dicts = [_row_to_dict(r) for r in rows]
    csv_str = export_csv(dicts, SHIFT_ATTENDANCE_COLUMNS)

    # Resolve shift name for filename
    shift_name = None
    if f.shift_id:
        st = db.query(ShiftTemplate.shift_name).filter(
            ShiftTemplate.id == f.shift_id,
            ShiftTemplate.tenant_id == f.tenant_id,
        ).scalar()
        shift_name = st

    tenant = db.query(Tenant).filter(Tenant.id == f.tenant_id).first()
    tenant_slug = (tenant.code if tenant else "unknown").lower()

    filename = generate_filename(
        "shift_attendance", tenant_slug,
        operating_date=f.operating_date,
        date_from=f.date_from, date_to=f.date_to,
        shift_id=shift_name, fmt="csv",
    )
    return csv_str, filename


def export_shift_attendance_xlsx(db: Session, f: ReportFilter) -> tuple[bytes, str]:
    """Export Report A as XLSX. Returns (xlsx_bytes, filename)."""
    rows = query_shift_attendance(db, f)
    dicts = [_row_to_dict(r) for r in rows]

    tenant = db.query(Tenant).filter(Tenant.id == f.tenant_id).first()
    site = None
    if f.site_id:
        site = db.query(Site).filter(Site.id == f.site_id, Site.tenant_id == f.tenant_id).first()
    shift_name = None
    if f.shift_id:
        shift_name = db.query(ShiftTemplate.shift_name).filter(
            ShiftTemplate.id == f.shift_id,
            ShiftTemplate.tenant_id == f.tenant_id,
        ).scalar()

    metadata = {
        "tenant_name": tenant.name if tenant else "",
        "tenant_code": tenant.code if tenant else "",
        "site_name": site.site_name if site else None,
        "operating_date": f.operating_date,
        "date_from": f.date_from,
        "date_to": f.date_to,
        "shift_name": shift_name,
        "timezone": tenant.timezone if tenant else "Asia/Jakarta",
        "generated_at": datetime.now(timezone.utc),
        "report_type": "Shift Attendance",
        "applied_filters": _build_applied_filters(f),
        "row_count": len(dicts),
    }

    tenant_slug = (tenant.code if tenant else "unknown").lower()
    filename = generate_filename(
        "shift_attendance", tenant_slug,
        operating_date=f.operating_date,
        date_from=f.date_from, date_to=f.date_to,
        shift_id=shift_name, fmt="xlsx",
    )

    xlsx_bytes = export_xlsx(dicts, SHIFT_ATTENDANCE_COLUMNS, "Shift_Attendance", metadata, filename)
    return xlsx_bytes, filename


def export_exception_decisions_csv(db: Session, f: ReportFilter) -> tuple[str, str]:
    """Export Report B as CSV. Returns (csv_string, filename)."""
    rows = query_exception_decisions(db, f)
    dicts = [_row_to_dict(r) for r in rows]
    csv_str = export_csv(dicts, EXCEPTION_DECISION_COLUMNS)

    tenant = db.query(Tenant).filter(Tenant.id == f.tenant_id).first()
    tenant_slug = (tenant.code if tenant else "unknown").lower()

    filename = generate_filename(
        "exceptions", tenant_slug,
        operating_date=f.operating_date,
        date_from=f.date_from, date_to=f.date_to,
        shift_id=None, fmt="csv",
    )
    return csv_str, filename


def export_exception_decisions_xlsx(db: Session, f: ReportFilter) -> tuple[bytes, str]:
    """Export Report B as XLSX. Returns (xlsx_bytes, filename)."""
    rows = query_exception_decisions(db, f)
    dicts = [_row_to_dict(r) for r in rows]

    tenant = db.query(Tenant).filter(Tenant.id == f.tenant_id).first()
    site = None
    if f.site_id:
        site = db.query(Site).filter(Site.id == f.site_id, Site.tenant_id == f.tenant_id).first()

    metadata = {
        "tenant_name": tenant.name if tenant else "",
        "tenant_code": tenant.code if tenant else "",
        "site_name": site.site_name if site else None,
        "operating_date": f.operating_date,
        "date_from": f.date_from,
        "date_to": f.date_to,
        "shift_name": None,
        "timezone": tenant.timezone if tenant else "Asia/Jakarta",
        "generated_at": datetime.now(timezone.utc),
        "report_type": "Exception & Decision",
        "applied_filters": _build_applied_filters(f),
        "row_count": len(dicts),
    }

    tenant_slug = (tenant.code if tenant else "unknown").lower()
    filename = generate_filename(
        "exceptions", tenant_slug,
        operating_date=f.operating_date,
        date_from=f.date_from, date_to=f.date_to,
        fmt="xlsx",
    )

    xlsx_bytes = export_xlsx(dicts, EXCEPTION_DECISION_COLUMNS, "Exception_Decision", metadata, filename)
    return xlsx_bytes, filename


def export_roster_vs_actual_csv(db: Session, f: ReportFilter) -> tuple[str, str]:
    """Export Report C as CSV. Returns (csv_string, filename)."""
    rows = query_roster_vs_actual(db, f)
    dicts = [_row_to_dict(r) for r in rows]
    csv_str = export_csv(dicts, ROSTER_VS_ACTUAL_COLUMNS)

    tenant = db.query(Tenant).filter(Tenant.id == f.tenant_id).first()
    shift_name = None
    if f.shift_id:
        shift_name = db.query(ShiftTemplate.shift_name).filter(
            ShiftTemplate.id == f.shift_id,
            ShiftTemplate.tenant_id == f.tenant_id,
        ).scalar()

    tenant_slug = (tenant.code if tenant else "unknown").lower()
    filename = generate_filename(
        "roster_vs_actual", tenant_slug,
        operating_date=f.operating_date,
        date_from=f.date_from, date_to=f.date_to,
        shift_id=shift_name, fmt="csv",
    )
    return csv_str, filename


def export_roster_vs_actual_xlsx(db: Session, f: ReportFilter) -> tuple[bytes, str]:
    """Export Report C as XLSX. Returns (xlsx_bytes, filename)."""
    rows = query_roster_vs_actual(db, f)
    dicts = [_row_to_dict(r) for r in rows]

    tenant = db.query(Tenant).filter(Tenant.id == f.tenant_id).first()
    site = None
    if f.site_id:
        site = db.query(Site).filter(Site.id == f.site_id, Site.tenant_id == f.tenant_id).first()
    shift_name = None
    if f.shift_id:
        shift_name = db.query(ShiftTemplate.shift_name).filter(
            ShiftTemplate.id == f.shift_id,
            ShiftTemplate.tenant_id == f.tenant_id,
        ).scalar()

    metadata = {
        "tenant_name": tenant.name if tenant else "",
        "tenant_code": tenant.code if tenant else "",
        "site_name": site.site_name if site else None,
        "operating_date": f.operating_date,
        "date_from": f.date_from,
        "date_to": f.date_to,
        "shift_name": shift_name,
        "timezone": tenant.timezone if tenant else "Asia/Jakarta",
        "generated_at": datetime.now(timezone.utc),
        "report_type": "Roster vs Actual Equipment",
        "applied_filters": _build_applied_filters(f),
        "row_count": len(dicts),
    }

    tenant_slug = (tenant.code if tenant else "unknown").lower()
    filename = generate_filename(
        "roster_vs_actual", tenant_slug,
        operating_date=f.operating_date,
        date_from=f.date_from, date_to=f.date_to,
        shift_id=shift_name, fmt="xlsx",
    )

    xlsx_bytes = export_xlsx(dicts, ROSTER_VS_ACTUAL_COLUMNS, "Roster_vs_Actual", metadata, filename)
    return xlsx_bytes, filename


# ── Public API aliases (used by reports.py router) ────────────

EXCEPTION_REPORT_COLUMNS = EXCEPTION_DECISION_COLUMNS


def get_shift_attendance_report(db: Session, f: ReportFilter) -> list[ShiftAttendanceRow]:
    """Alias for query_shift_attendance."""
    return query_shift_attendance(db, f)


def get_exception_report(db: Session, f: ReportFilter) -> list[ExceptionDecisionRow]:
    """Alias for query_exception_decisions."""
    return query_exception_decisions(db, f)


def get_roster_vs_actual_report(db: Session, f: ReportFilter) -> list[RosterVsActualRow]:
    """Alias for query_roster_vs_actual."""
    return query_roster_vs_actual(db, f)


def shift_attendance_to_dicts(rows: list[ShiftAttendanceRow]) -> list[dict]:
    """Convert ShiftAttendanceRow list to dicts."""
    return [_row_to_dict(r) for r in rows]


def exception_report_to_dicts(rows: list[ExceptionDecisionRow]) -> list[dict]:
    """Convert ExceptionDecisionRow list to dicts."""
    return [_row_to_dict(r) for r in rows]


def roster_vs_actual_to_dicts(rows: list[RosterVsActualRow]) -> list[dict]:
    """Convert RosterVsActualRow list to dicts."""
    return [_row_to_dict(r) for r in rows]


def build_report_metadata(
    db: Session, ctx, report_type: str, f: ReportFilter, row_count: int,
) -> ReportMetadata:
    """Build ReportMetadata for export."""
    tenant = db.query(Tenant).filter(Tenant.id == f.tenant_id).first()
    site = None
    if f.site_id:
        site = db.query(Site).filter(Site.id == f.site_id, Site.tenant_id == f.tenant_id).first()
    shift_name = None
    if f.shift_id:
        shift_name = db.query(ShiftTemplate.shift_name).filter(
            ShiftTemplate.id == f.shift_id,
            ShiftTemplate.tenant_id == f.tenant_id,
        ).scalar()

    return ReportMetadata(
        tenant_name=tenant.name if tenant else "",
        tenant_code=tenant.code if tenant else "",
        site_name=site.site_name if site else None,
        operating_date=f.operating_date,
        date_from=f.date_from,
        date_to=f.date_to,
        shift_name=shift_name,
        timezone=tenant.timezone if tenant else "Asia/Jakarta",
        generated_at=datetime.now(timezone.utc),
        report_type=report_type,
        applied_filters=_build_applied_filters(f),
        row_count=row_count,
    )
